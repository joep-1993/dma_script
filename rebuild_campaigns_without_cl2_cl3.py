"""
Rebuild campaigns with shop exclusions (UPDATED VERSION).

This script:
1. Groups rows by campaign (cat_uitsluiten + diepste_cat_id + custom_label_1)
2. Collects all shops to exclude for each campaign
3. Removes the entire listing tree
4. Rebuilds it with CL0, CL1, and multiple CL3 shop exclusions
"""
from openpyxl import load_workbook
from collections import defaultdict
from campaign_processor import (
    initialize_google_ads_client,
    CUSTOMER_ID,
    DEFAULT_BID_MICROS
)
from google_ads_helpers import (
    safe_remove_entire_listing_tree,
    create_listing_group_subdivision,
    create_listing_group_unit_biddable
)

# File and sheet configuration
EXCEL_FILE = "/mnt/c/Users/JoepvanSchagen/Downloads/dma_script_ivor.xlsx"
SHEET_NAME = "uitsluiten"

# Column indices (0-based) - CORRECTED
COL_SHOP_NAME = 0       # A: Shop name
COL_SHOP_ID = 1         # B: Shop ID
COL_CAT_UITSLUITEN = 2  # C: cat_uitsluiten
COL_DIEPSTE_CAT_ID = 3  # D: Diepste cat ID
COL_CUSTOM_LABEL_1 = 4  # E: custom label 1
COL_RESULT = 5          # F: result


def get_campaign_and_ad_group_by_pattern(client, customer_id, campaign_pattern):
    """Find campaign and ad group by exact campaign name pattern"""
    ga_service = client.get_service("GoogleAdsService")
    escaped_pattern = campaign_pattern.replace("'", "\\'")

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            ad_group.id,
            ad_group.name
        FROM ad_group
        WHERE campaign.name = '{escaped_pattern}'
        AND campaign.status != REMOVED
        AND ad_group.status != REMOVED
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)
        for row in response:
            return {
                'campaign': {
                    'id': row.campaign.id,
                    'name': row.campaign.name
                },
                'ad_group': {
                    'id': row.ad_group.id,
                    'name': row.ad_group.name
                }
            }
    except Exception as e:
        return None

    return None


def rebuild_tree_with_shop_exclusions(client, customer_id, ad_group_id, diepste_cat_id, custom_label_1, shop_names, default_bid_micros=DEFAULT_BID_MICROS):
    """
    Rebuild listing tree with CL0, CL1, and multiple CL3 shop exclusions.

    Tree structure:
    ROOT (subdivision)
    ├─ CL0 = diepste_cat_id (subdivision)
    │  ├─ CL1 = custom_label_1 (subdivision)
    │  │  ├─ CL3 = shop1 (unit, negative)
    │  │  ├─ CL3 = shop2 (unit, negative)
    │  │  ├─ CL3 = shop3 (unit, negative)
    │  │  └─ CL3 OTHERS (unit, positive with bid)
    │  └─ CL1 OTHERS (unit, negative)
    └─ CL0 OTHERS (unit, negative)

    Args:
        shop_names: List of shop names to exclude (CL3)
    """
    # Step 1: Remove entire existing tree
    safe_remove_entire_listing_tree(client, customer_id, str(ad_group_id))

    # Step 2: Build new tree with multiple operations
    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create ROOT, CL0, and their OTHERS cases
    ops1 = []

    # ROOT
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # CL0 subdivision (diepste_cat_id)
    dim_cl0 = client.get_type("ListingDimensionInfo")
    dim_cl0.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
    dim_cl0.product_custom_attribute.value = str(diepste_cat_id)

    cl0_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl0
    )
    cl0_subdivision_tmp = cl0_subdivision_op.create.resource_name
    ops1.append(cl0_subdivision_op)

    # CL0 OTHERS (negative - under root)
    dim_cl0_others = client.get_type("ListingDimensionInfo")
    dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl0_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    try:
        response1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
        cl0_actual = response1.results[1].resource_name  # Get actual CL0 resource name
    except Exception as e:
        raise Exception(f"Error creating ROOT and CL0: {e}")

    # MUTATE 2: Create CL1 subdivision and its OTHERS
    ops2 = []

    # CL1 subdivision (custom_label_1 - under CL0)
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    dim_cl1.product_custom_attribute.value = str(custom_label_1)

    cl1_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=cl0_actual,
        listing_dimension_info=dim_cl1
    )
    cl1_subdivision_tmp = cl1_subdivision_op.create.resource_name
    ops2.append(cl1_subdivision_op)

    # CL1 OTHERS (negative - under CL0)
    dim_cl1_others = client.get_type("ListingDimensionInfo")
    dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    ops2.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=cl0_actual,
            listing_dimension_info=dim_cl1_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # Execute second mutate
    try:
        response2 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
        cl1_actual = response2.results[0].resource_name  # Get actual CL1 resource name
    except Exception as e:
        raise Exception(f"Error creating CL1: {e}")

    # MUTATE 3: Create CL3 shop exclusions and CL3 OTHERS
    ops3 = []

    # Add each shop as a negative CL3 unit
    for shop_name in shop_names:
        dim_cl3_shop = client.get_type("ListingDimensionInfo")
        dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
        dim_cl3_shop.product_custom_attribute.value = str(shop_name)

        ops3.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=cl1_actual,
                listing_dimension_info=dim_cl3_shop,
                targeting_negative=True,  # NEGATIVE = exclude this shop
                cpc_bid_micros=None
            )
        )

    # CL3 OTHERS (positive with bid - under CL1)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    ops3.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=cl1_actual,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=False,  # POSITIVE = target all other shops
            cpc_bid_micros=default_bid_micros
        )
    )

    # Execute third mutate
    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops3)
        return True
    except Exception as e:
        raise Exception(f"Error adding shop exclusions: {e}")


def main():
    print("="*70)
    print("REBUILDING CAMPAIGNS WITH SHOP EXCLUSIONS")
    print("="*70)
    print(f"\nFile: {EXCEL_FILE}")
    print(f"Sheet: {SHEET_NAME}")

    # Load workbook
    print(f"\nLoading workbook...")
    wb = load_workbook(EXCEL_FILE)

    try:
        sheet = wb[SHEET_NAME]
    except KeyError:
        print(f"❌ Sheet '{SHEET_NAME}' not found!")
        return

    print(f"✅ Loaded sheet: {SHEET_NAME}")

    # Initialize Google Ads client
    print("\nInitializing Google Ads client...")
    client = initialize_google_ads_client()

    # Step 1: Group rows by campaign and collect all shops to exclude
    print("\n" + "="*70)
    print("STEP 1: GROUPING ROWS BY CAMPAIGN")
    print("="*70)

    # Group by (cat_uitsluiten, diepste_cat_id, custom_label_1)
    campaign_groups = defaultdict(lambda: {'rows': [], 'shops': set()})

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Skip rows already processed
        result_value = row[COL_RESULT].value
        if result_value is not None and result_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value
        cat_uitsluiten = row[COL_CAT_UITSLUITEN].value
        diepste_cat_id = row[COL_DIEPSTE_CAT_ID].value
        custom_label_1 = row[COL_CUSTOM_LABEL_1].value

        if not shop_name or not cat_uitsluiten or not diepste_cat_id or not custom_label_1:
            continue

        # Group key
        group_key = (cat_uitsluiten, str(diepste_cat_id), str(custom_label_1))

        # Add row and shop to group
        campaign_groups[group_key]['rows'].append({
            'idx': idx,
            'row_obj': row
        })
        campaign_groups[group_key]['shops'].add(str(shop_name))

    print(f"Found {len(campaign_groups)} campaign group(s) to process")
    print(f"Total rows: {sum(len(g['rows']) for g in campaign_groups.values())}\n")

    # Step 2: Process each group
    print("="*70)
    print("STEP 2: PROCESSING CAMPAIGN GROUPS")
    print("="*70)

    success_count = 0
    fail_count = 0
    groups_processed = 0

    for i, (group_key, group_data) in enumerate(campaign_groups.items(), 1):
        cat_uitsluiten, diepste_cat_id, custom_label_1 = group_key
        rows = group_data['rows']
        shops = sorted(group_data['shops'])  # Sort for consistent output

        campaign_pattern = f"PLA/{cat_uitsluiten}_{custom_label_1}"

        print(f"\n{'─'*70}")
        print(f"GROUP {i}/{len(campaign_groups)}: {campaign_pattern}")
        print(f"{'─'*70}")
        print(f"   Rows in group: {len(rows)}")
        print(f"   Shops to exclude: {len(shops)}")
        print(f"   Shop names: {', '.join(shops)}")

        try:
            # Find campaign
            result = get_campaign_and_ad_group_by_pattern(client, CUSTOMER_ID, campaign_pattern)

            if not result:
                print(f"   ❌ Campaign not found")
                # Mark all rows in group as NOT_FOUND
                for row_info in rows:
                    row_info['row_obj'][COL_RESULT].value = "NOT_FOUND"
                    fail_count += 1
                continue

            print(f"   ✅ Found: Campaign ID {result['campaign']['id']}, Ad Group ID {result['ad_group']['id']}")
            print(f"   Rebuilding with: CL0={diepste_cat_id}, CL1={custom_label_1}")
            print(f"   Excluding {len(shops)} shop(s)...")

            # Rebuild tree with all shop exclusions
            rebuild_tree_with_shop_exclusions(
                client,
                CUSTOMER_ID,
                result['ad_group']['id'],
                diepste_cat_id,
                custom_label_1,
                shops  # Pass all shops for this campaign
            )

            # Mark all rows in group as SUCCESS
            for row_info in rows:
                row_info['row_obj'][COL_RESULT].value = True
                success_count += 1

            groups_processed += 1
            print(f"   ✅ SUCCESS - Tree rebuilt with {len(shops)} shop exclusion(s)")

        except Exception as e:
            print(f"   ❌ ERROR: {e}")
            # Mark all rows in group as ERROR
            error_msg = f"ERROR: {str(e)[:100]}"
            for row_info in rows:
                row_info['row_obj'][COL_RESULT].value = error_msg
                fail_count += 1

        # Save every 10 groups
        if i % 10 == 0:
            wb.save(EXCEL_FILE)
            print(f"\n   💾 Progress saved ({i}/{len(campaign_groups)} groups processed)")

    # Final save
    wb.save(EXCEL_FILE)

    # Summary
    print("\n" + "="*70)
    print("SUMMARY")
    print("="*70)
    print(f"Total campaign groups processed: {len(campaign_groups)}")
    print(f"✅ Groups successful: {groups_processed}")
    print(f"❌ Groups failed: {len(campaign_groups) - groups_processed}")
    print(f"✅ Total rows marked success: {success_count}")
    print(f"❌ Total rows marked failed: {fail_count}")
    print("="*70)


if __name__ == "__main__":
    main()
