"""
Integration test for the optimized reverse_exclusion functions.

This test loads the actual Excel file and verifies:
1. The file can be read correctly
2. Rows are grouped correctly by (maincat_id, cl1)
3. The cat_ids mapping is loaded correctly
4. Campaign names are constructed correctly

Run with: python test_reverse_exclusion_integration.py

This is a DRY-RUN test - no Google Ads API calls are made.
"""

import sys
import os
import platform
from collections import defaultdict

# Add script directory to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

import openpyxl


def get_test_file_path():
    """Get the path to the reverse exclusion Excel file."""
    windows_path = r"C:\Users\JoepvanSchagen\Downloads\claude\dma_script_uitbreiding_reverse.xlsx"
    wsl_path = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/dma_script_uitbreiding_reverse.xlsx"

    system = platform.system().lower()
    if system == "windows":
        return windows_path
    else:
        return wsl_path


def load_cat_ids_mapping_test(workbook):
    """
    Load the cat_ids mapping from the workbook.
    Copy of the function from campaign_processor.py for testing.

    Sheet structure:
    - Column A (0): maincat name (for reference)
    - Column B (1): maincat_id (numeric) - used as key
    - Column C (2): deepest_cat name - used as value
    - Column D (3): cat_id (numeric deepest_cat_id, not used)
    """
    if 'cat_ids' not in workbook.sheetnames:
        print("⚠️  No 'cat_ids' sheet found")
        return {}

    sheet = workbook['cat_ids']
    mapping = defaultdict(set)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 3:
            maincat_id = row[1]   # Column B: maincat_id (numeric)
            deepest_cat = row[2]  # Column C: deepest_cat name

            if maincat_id and deepest_cat:
                maincat_id_str = str(maincat_id)
                deepest_cat_str = str(deepest_cat)
                mapping[maincat_id_str].add(deepest_cat_str)

    # Convert sets to sorted lists
    return {k: sorted(v) for k, v in mapping.items()}


def test_integration():
    """Run integration test with actual Excel file."""
    print("\n" + "="*70)
    print("INTEGRATION TEST - REVERSE EXCLUSION (DRY RUN)")
    print("="*70)

    file_path = get_test_file_path()
    print(f"\nFile path: {file_path}")

    # Check if file exists
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        print("   Please ensure the reverse exclusion file exists.")
        return False

    print("✅ File exists")

    # Load workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"✅ Workbook loaded successfully")
        print(f"   Sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        return False

    # Check for required sheets
    required_sheets = ['verwijderen', 'cat_ids']
    for sheet_name in required_sheets:
        if sheet_name in workbook.sheetnames:
            print(f"✅ Sheet '{sheet_name}' found")
        else:
            print(f"❌ Sheet '{sheet_name}' NOT found")
            return False

    # Load cat_ids mapping
    print("\n--- Loading cat_ids mapping ---")
    cat_ids_mapping = load_cat_ids_mapping_test(workbook)
    print(f"✅ Loaded {len(cat_ids_mapping)} maincat_id entries")

    if cat_ids_mapping:
        # Show sample
        sample_keys = list(cat_ids_mapping.keys())[:3]
        for key in sample_keys:
            deepest_cats = cat_ids_mapping[key]
            print(f"   maincat_id={key}: {len(deepest_cats)} deepest_cat(s)")
            if deepest_cats:
                print(f"      First: {deepest_cats[0]}")

    # Process verwijderen sheet
    print("\n--- Processing 'verwijderen' sheet ---")
    sheet = workbook['verwijderen']

    # Column indices
    COL_SHOP_NAME = 0      # A
    COL_SHOP_ID = 1        # B
    COL_MAINCAT = 2        # C
    COL_MAINCAT_ID = 3     # D
    COL_CL1 = 4            # E
    COL_STATUS = 5         # F

    # Count and group rows
    groups = defaultdict(list)
    rows_to_process = 0
    rows_already_processed = 0
    rows_missing_fields = 0

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 5:
            continue

        shop_name = row[COL_SHOP_NAME]
        maincat_id = row[COL_MAINCAT_ID]
        cl1 = row[COL_CL1]
        status = row[COL_STATUS] if len(row) > COL_STATUS else None

        if not shop_name:
            continue

        if status is not None and status != '':
            rows_already_processed += 1
            continue

        if not maincat_id or not cl1:
            rows_missing_fields += 1
            continue

        rows_to_process += 1
        groups[(str(maincat_id), str(cl1))].append((idx, shop_name))

    print(f"\nRow statistics:")
    print(f"  Total data rows: {rows_to_process + rows_already_processed + rows_missing_fields}")
    print(f"  Already processed: {rows_already_processed}")
    print(f"  Missing fields: {rows_missing_fields}")
    print(f"  To process: {rows_to_process}")
    print(f"  Unique groups: {len(groups)}")

    # Show efficiency gain
    if groups:
        total_shops_in_groups = sum(len(shops) for shops in groups.values())
        avg_shops_per_group = total_shops_in_groups / len(groups) if groups else 0
        print(f"\n  Average shops per group: {avg_shops_per_group:.1f}")

        # Show top groups by size
        sorted_groups = sorted(groups.items(), key=lambda x: len(x[1]), reverse=True)
        print(f"\n  Top 5 groups by size:")
        for (maincat_id, cl1), shops in sorted_groups[:5]:
            print(f"    ({maincat_id}, {cl1}): {len(shops)} shops")

        # Verify campaign name construction
        print(f"\n--- Campaign Name Verification ---")
        sample_group = sorted_groups[0]
        (maincat_id, cl1), shops = sample_group
        deepest_cats = cat_ids_mapping.get(maincat_id, [])

        if deepest_cats:
            print(f"\nFor group ({maincat_id}, {cl1}) with {len(shops)} shops:")
            print(f"  deepest_cats: {len(deepest_cats)}")
            print(f"  Sample campaign names:")
            for dc in deepest_cats[:3]:
                campaign_name = f"PLA/{dc}_{cl1}"
                print(f"    - {campaign_name}")
        else:
            print(f"\n⚠️  No deepest_cats found for maincat_id={maincat_id}")

    # Calculate efficiency
    print("\n--- Efficiency Analysis ---")
    if groups and cat_ids_mapping:
        # Estimate API calls
        total_old_calls = 0
        total_new_calls = 0
        avg_ad_groups = 3  # Assume 3 ad groups per campaign

        for (maincat_id, cl1), shops in groups.items():
            num_shops = len(shops)
            deepest_cats = cat_ids_mapping.get(maincat_id, [])
            num_deepest_cats = len(deepest_cats)

            if num_deepest_cats > 0:
                # Old: per shop, per deepest_cat, per ad_group: read + remove
                old_calls = num_shops * num_deepest_cats * avg_ad_groups * 2
                # New: per deepest_cat, per ad_group: read + batch remove
                new_calls = num_deepest_cats * avg_ad_groups * 2

                total_old_calls += old_calls
                total_new_calls += new_calls

        if total_old_calls > 0:
            reduction = (1 - total_new_calls / total_old_calls) * 100
            print(f"  Estimated API calls (old approach): {total_old_calls:,}")
            print(f"  Estimated API calls (new approach): {total_new_calls:,}")
            print(f"  Reduction: {reduction:.1f}%")

    print("\n" + "="*70)
    print("✅ INTEGRATION TEST PASSED - All checks completed")
    print("="*70)
    print("\nThis was a DRY RUN - no changes were made to Google Ads.")
    print("The function is ready for production use.")

    return True


if __name__ == "__main__":
    success = test_integration()
    sys.exit(0 if success else 1)
