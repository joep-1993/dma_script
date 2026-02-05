"""
DMA Shop Campaigns Processor

This script processes Excel files with shop campaign data and updates Google Ads
listing trees with custom label 3 targeting (shop name).

Usage:
    python campaign_processor.py

Configuration:
    - Excel file path: EXCEL_FILE_PATH constant below
    - Customer ID: CUSTOMER_ID constant below
    - Google Ads credentials: google-ads.yaml in the same directory or set via environment
"""

import sys
import os
import time
import platform
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional, Dict, Any
from google.ads.googleads.client import GoogleAdsClient
from google.ads.googleads.errors import GoogleAdsException
import openpyxl
from openpyxl import load_workbook
from dotenv import load_dotenv
import shutil
from datetime import datetime

# Load environment variables
load_dotenv()

# Add script directory to Python path for imports
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

# Import helper functions (add your existing helper functions to google_ads_helpers.py)
try:
    from google_ads_helpers import (
        safe_remove_entire_listing_tree,
        create_listing_group_subdivision,
        create_listing_group_unit_biddable,
        add_standard_shopping_campaign,
        add_shopping_ad_group,
        add_shopping_product_ad,
        enable_negative_list_for_campaign,
    )
except ImportError as e:
    print(f"⚠️  Warning: Could not import helper functions from google_ads_helpers.py")
    print(f"   Error: {e}")
    print(f"   Script directory: {SCRIPT_DIR}")
    print(f"   Please ensure google_ads_helpers.py is in the same directory as this script")

# ============================================================================
# CONFIGURATION
# ============================================================================

CUSTOMER_ID = "3800751597"
MCC_ACCOUNT_ID = "3011145605"  # MCC account where bid strategies are stored
DEFAULT_BID_MICROS = 200_000  # €0.20

# Negative keyword list to add to all created campaigns
NEGATIVE_LIST_NAME = "DMA negatives"

# Bid strategy mapping based on custom label 1
BID_STRATEGY_MAPPING = {
    'a': 'DMA: Elektronica shops A - 0,25',
    'b': 'DMA: Elektronica shops B - 0,21',
    'c': 'DMA: Elektronica shops C - 0,17'
}

# Auto-detect Excel file path based on operating system
def get_excel_path():
    """
    Automatically detect the correct Excel file path based on OS.

    Returns:
        str: Path to Excel file (WSL format for Linux, Windows format for Windows)
    """
    # Base Windows path
    windows_path = "c:/Users/JoepvanSchagen/Downloads/Python/scripts_def/DMA+/dma_script_uitbreiding.xlsx"
    wsl_path = "/mnt/c/Users/JoepvanSchagen/Downloads/Python/scripts_def/DMA+/dma_script_uitbreiding.xlsx"

    system = platform.system().lower()

    if system == "windows":
        # Running on native Windows (PyCharm on Windows)
        return windows_path
    elif system == "linux":
        # Check if running on WSL
        if os.path.exists("/proc/version"):
            with open("/proc/version", "r") as f:
                if "microsoft" in f.read().lower():
                    # Running on WSL
                    return wsl_path
        # Running on native Linux - try WSL path first, fall back to Windows path
        if os.path.exists(wsl_path):
            return wsl_path
        return windows_path
    else:
        # Default to Windows path for other systems (macOS, etc.)
        return windows_path

EXCEL_FILE_PATH = get_excel_path()

def get_reverse_exclusion_path():
    """
    Get the path to the reverse exclusion Excel file.

    Returns:
        str: Path to reverse exclusion Excel file (WSL format for Linux, Windows format for Windows)
    """
    windows_path = "C:/Users/JoepvanSchagen/Downloads/claude/dma_script_uitbreiding_reverse.xlsx"
    wsl_path = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/dma_script_uitbreiding_reverse.xlsx"

    system = platform.system().lower()

    if system == "windows":
        return windows_path
    elif system == "linux":
        if os.path.exists("/proc/version"):
            with open("/proc/version", "r") as f:
                if "microsoft" in f.read().lower():
                    return wsl_path
        if os.path.exists(wsl_path):
            return wsl_path
        return windows_path
    else:
        return windows_path

REVERSE_EXCLUSION_FILE_PATH = get_reverse_exclusion_path()

# Sheet names
SHEET_INCLUSION = "toevoegen"  # Inclusion sheet
SHEET_EXCLUSION = "uitsluiten"  # Exclusion sheet
SHEET_REVERSE_INCLUSION = "toevoegen"  # Reverse inclusion sheet (remove ad groups)
SHEET_ENABLE_INCLUSION = "adgroups_heractiveren"  # Enable inclusion sheet (enable ad groups)

# Column indices (0-based) - INCLUSION SHEET (toevoegen) - NEW STRUCTURE (v2)
COL_CAMPAIGN_NAME = 0  # Column A: campaign_name
COL_AD_GROUP_NAME = 1  # Column B: ad group_name (also used as shop_name for CL3)
COL_SHOP_ID = 2        # Column C: Shop ID
COL_MAINCAT = 3        # Column D: maincat
COL_MAINCAT_ID = 4     # Column E: maincat_id (used as CL4)
COL_CUSTOM_LABEL_1 = 5 # Column F: custom label 1
COL_BUDGET = 6         # Column G: budget
COL_STATUS = 7         # Column H: result (TRUE/FALSE)
COL_ERROR = 8          # Column I: Error message (when status is FALSE)

# Column indices (0-based) - INCLUSION SHEET (toevoegen) - LEGACY STRUCTURE
COL_LEGACY_SHOP_NAME = 0      # Column A: Shop name
COL_LEGACY_SHOP_ID = 1        # Column B: Shop ID
COL_LEGACY_MAINCAT = 2        # Column C: maincat
COL_LEGACY_MAINCAT_ID = 3     # Column D: maincat_id
COL_LEGACY_CUSTOM_LABEL_1 = 4 # Column E: custom label 1
COL_LEGACY_BUDGET = 5         # Column F: budget
COL_LEGACY_STATUS = 6         # Column G: Status (TRUE/FALSE)
COL_LEGACY_ERROR = 7          # Column H: Error message (when status is FALSE)

# Column indices (0-based) - EXCLUSION SHEET (uitsluiten) - NEW STRUCTURE
COL_EX_SHOP_NAME = 0      # Column A: Shop name
COL_EX_SHOP_ID = 1        # Column B: Shop ID
COL_EX_MAINCAT = 2        # Column C: maincat
COL_EX_MAINCAT_ID = 3     # Column D: maincat_id
COL_EX_CUSTOM_LABEL_1 = 4 # Column E: custom label 1
COL_EX_STATUS = 5         # Column F: result (TRUE/FALSE)
COL_EX_ERROR = 6          # Column G: Error message (when status is FALSE)

# Column indices (0-based) - CAT_IDS SHEET (category mappings)
COL_CAT_MAINCAT = 0       # Column A: maincat
COL_CAT_MAINCAT_ID = 1    # Column B: maincat_id
COL_CAT_DEEPEST_CAT = 2   # Column C: deepest_cat
COL_CAT_CAT_ID = 3        # Column D: cat_id

# Sheet names
SHEET_CAT_IDS = "cat_ids"

# Column indices (0-based) - UITBREIDING SHEET (extension/expansion)
COL_UIT_SHOP_NAME = 0      # Column A: Shop name
COL_UIT_SHOP_ID = 1        # Column B: Shop ID (not used)
COL_UIT_MAINCAT = 2        # Column C: maincat (category name)
COL_UIT_MAINCAT_ID = 3     # Column D: maincat_id (used as CL4)
COL_UIT_CUSTOM_LABEL_1 = 4 # Column E: custom label 1 (a/b/c)
COL_UIT_BUDGET = 5         # Column F: budget
COL_UIT_STATUS = 6         # Column G: result (TRUE/FALSE)
COL_UIT_ERROR = 7          # Column H: Error message (when status is FALSE)

# Sheet name for uitbreiding
SHEET_UITBREIDING = "toevoegen"  # Using same sheet name as inclusion


# ============================================================================
# GOOGLE ADS CLIENT INITIALIZATION
# ============================================================================

def load_google_oauth_from_env():
    """
    Load Google OAuth credentials from environment variables.
    Uses the same env vars as the working GSD-campaigns script.
    """
    client_id = os.getenv("GOOGLE_CLIENT_ID")
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET")
    missing = []
    if not client_id:
        missing.append("GOOGLE_CLIENT_ID")
    if not client_secret:
        missing.append("GOOGLE_CLIENT_SECRET")
    if missing:
        raise RuntimeError(
            f"Environment variables missing: {', '.join(missing)}.\n"
            "Set them in Windows with: setx GOOGLE_CLIENT_ID \"...\" and setx GOOGLE_CLIENT_SECRET \"...\"\n"
            "Or temporarily with: set GOOGLE_CLIENT_ID=... & set GOOGLE_CLIENT_SECRET=..."
        )
    return client_id, client_secret


def initialize_google_ads_client():
    """
    Initialize Google Ads API client.

    Uses the same authentication approach as 'create GSD-campaigns WB.py':
    - Loads all credentials from environment variables

    Returns:
        GoogleAdsClient: Initialized client

    Required environment variables:
        GOOGLE_ADS_REFRESH_TOKEN: OAuth refresh token
        GOOGLE_ADS_DEVELOPER_TOKEN: Developer token
        GOOGLE_ADS_LOGIN_CUSTOMER_ID: MCC account ID (default: 3011145605)
        GOOGLE_CLIENT_ID: OAuth client ID
        GOOGLE_CLIENT_SECRET: OAuth client secret
    """
    try:
        # Load all credentials from environment variables
        print("Loading Google Ads credentials from environment variables...")
        client_id, client_secret = load_google_oauth_from_env()

        refresh_token = os.getenv("GOOGLE_ADS_REFRESH_TOKEN")
        developer_token = os.getenv("GOOGLE_ADS_DEVELOPER_TOKEN")
        login_customer_id = os.getenv("GOOGLE_ADS_LOGIN_CUSTOMER_ID", "3011145605")

        if not refresh_token:
            raise RuntimeError("GOOGLE_ADS_REFRESH_TOKEN environment variable not set")
        if not developer_token:
            raise RuntimeError("GOOGLE_ADS_DEVELOPER_TOKEN environment variable not set")

        credentials = {
            "developer_token": developer_token,
            "refresh_token": refresh_token,
            "client_id": client_id,
            "client_secret": client_secret,
            "login_customer_id": login_customer_id,
            "use_proto_plus": True
        }

        client = GoogleAdsClient.load_from_dict(credentials)
        print("✅ Google Ads client initialized successfully")

        return client
    except RuntimeError as e:
        print(f"❌ Error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error initializing Google Ads client: {e}")
        print("   Make sure your environment variables are set:")
        print("   - GOOGLE_CLIENT_ID")
        print("   - GOOGLE_CLIENT_SECRET")
        sys.exit(1)


# ============================================================================
# BID STRATEGY RETRIEVAL
# ============================================================================

def get_bid_strategy_by_name(
    client: GoogleAdsClient,
    customer_id: str,
    strategy_name: str
) -> Optional[str]:
    """
    Retrieve portfolio bid strategy by name.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        strategy_name: Bid strategy name to search for

    Returns:
        Bid strategy resource name or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    # Escape single quotes in strategy name for GAQL (replace ' with \')
    escaped_strategy_name = strategy_name.replace("'", "\\'")

    query = f"""
        SELECT
            bidding_strategy.id,
            bidding_strategy.name,
            bidding_strategy.resource_name
        FROM bidding_strategy
        WHERE bidding_strategy.name = '{escaped_strategy_name}'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            print(f"   📊 Found bid strategy: {row.bidding_strategy.name} (ID: {row.bidding_strategy.id})")
            return row.bidding_strategy.resource_name

        print(f"   ⚠️  Bid strategy '{strategy_name}' not found")
        return None

    except Exception as e:
        print(f"   ❌ Error searching for bid strategy '{strategy_name}': {e}")
        return None


# ============================================================================
# CAMPAIGN AND AD GROUP RETRIEVAL
# ============================================================================

def get_campaign_by_name_pattern(
    client: GoogleAdsClient,
    customer_id: str,
    name_pattern: str
) -> Optional[Dict[str, Any]]:
    """
    Retrieve campaign by name pattern.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        name_pattern: Campaign name pattern (e.g., "PLA/Electronics_A")

    Returns:
        Dict with campaign info (id, name, resource_name) or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    # Escape single quotes in name pattern for GAQL (replace ' with \')
    escaped_name_pattern = name_pattern.replace("'", "\\'")

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            campaign.status
        FROM campaign
        WHERE campaign.name LIKE '%{escaped_name_pattern}%'
            AND campaign.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            campaign = row.campaign
            return {
                'id': campaign.id,
                'name': campaign.name,
                'resource_name': campaign.resource_name,
                'status': campaign.status.name
            }

        return None

    except GoogleAdsException as e:
        print(f"❌ Error searching for campaign '{name_pattern}': {e}")
        return None


def get_ad_group_from_campaign(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_id: int
) -> Optional[Dict[str, Any]]:
    """
    Retrieve the first active ad group from a campaign.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_id: Campaign ID

    Returns:
        Dict with ad group info (id, name, resource_name) or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    query = f"""
        SELECT
            ad_group.id,
            ad_group.name,
            ad_group.resource_name,
            ad_group.status
        FROM ad_group
        WHERE ad_group.campaign = 'customers/{customer_id}/campaigns/{campaign_id}'
            AND ad_group.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            ad_group = row.ad_group
            return {
                'id': ad_group.id,
                'name': ad_group.name,
                'resource_name': ad_group.resource_name,
                'status': ad_group.status.name
            }

        return None

    except GoogleAdsException as e:
        print(f"❌ Error retrieving ad group for campaign {campaign_id}: {e}")
        return None


def get_campaign_and_ad_group_by_pattern(
    client: GoogleAdsClient,
    customer_id: str,
    name_pattern: str
) -> Optional[Dict[str, Any]]:
    """
    Retrieve campaign AND ad group by campaign name pattern in a single query.
    This is more efficient than making two separate API calls.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        name_pattern: Campaign name pattern (e.g., "PLA/Electronics_A")

    Returns:
        Dict with campaign and ad_group info:
        {
            'campaign': {'id': ..., 'name': ..., 'resource_name': ..., 'status': ...},
            'ad_group': {'id': ..., 'name': ..., 'resource_name': ..., 'status': ...}
        }
        or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    # Escape single quotes in name pattern for GAQL (replace ' with \')
    escaped_name_pattern = name_pattern.replace("'", "\\'")

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            campaign.status,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name,
            ad_group.status
        FROM ad_group
        WHERE campaign.name LIKE '%{escaped_name_pattern}%'
            AND campaign.status != 'REMOVED'
            AND ad_group.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            return {
                'campaign': {
                    'id': row.campaign.id,
                    'name': row.campaign.name,
                    'resource_name': row.campaign.resource_name,
                    'status': row.campaign.status.name
                },
                'ad_group': {
                    'id': row.ad_group.id,
                    'name': row.ad_group.name,
                    'resource_name': row.ad_group.resource_name,
                    'status': row.ad_group.status.name
                }
            }

        return None

    except GoogleAdsException as e:
        print(f"❌ Error searching for campaign+ad group '{name_pattern}': {e}")
        return None


# ============================================================================
# LISTING TREE REBUILD FUNCTIONS (Custom Label 3 Targeting)
# ============================================================================

def rebuild_tree_with_custom_label_3_inclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: int,
    shop_name: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Rebuild listing tree to TARGET (include) a specific shop name via custom label 3.

    Structure:
    Root SUBDIVISION
    ├─ Custom Label 3 = shop_name [POSITIVE, biddable] → Target this shop
    └─ Custom Label 3 OTHERS [NEGATIVE] → Exclude all other shops

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to target (custom label 3 value)
        default_bid_micros: Bid amount in micros
    """
    print(f"   Rebuilding tree to TARGET shop '{shop_name}' (custom label 3)")

    # Remove existing tree
    safe_remove_entire_listing_tree(client, customer_id, str(ad_group_id))
    time.sleep(0.5)

    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create root SUBDIVISION + Custom Label 3 OTHERS (negative)
    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # 2. Custom Label 3 OTHERS (negative - blocks all other shops)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3  # INDEX3 = Custom Label 3
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,  # NEGATIVE - blocks everything else
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    root_actual = resp1.results[0].resource_name
    time.sleep(0.5)

    # MUTATE 2: Add specific shop name as POSITIVE unit
    ops2 = []

    dim_shop = client.get_type("ListingDimensionInfo")
    dim_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3  # INDEX3 = Custom Label 3
    dim_shop.product_custom_attribute.value = shop_name

    ops2.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=root_actual,
            listing_dimension_info=dim_shop,
            targeting_negative=False,  # POSITIVE targeting
            cpc_bid_micros=default_bid_micros
        )
    )

    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    print(f"   ✅ Tree rebuilt: ONLY targeting shop '{shop_name}'")


def rebuild_tree_with_custom_label_3_exclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: int,
    shop_name: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Rebuild listing tree to EXCLUDE a specific shop name via custom label 3.

    Following the pattern from rebuild_tree_with_label_and_item_ids in example_functions.txt:
    1. Read existing tree structure
    2. Collect ALL custom label structures (CL0, CL1, etc.) EXCEPT CL3
    3. Rebuild tree preserving those structures
    4. Add CL3 exclusion

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to exclude (custom label 3 value)
        default_bid_micros: Bid amount in micros
    """
    print(f"   Rebuilding tree to EXCLUDE shop '{shop_name}' (custom label 3)")

    # Step 1: Read existing tree structure
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"   ❌ Error reading existing tree: {e}")
        raise  # Re-raise exception so calling code can handle it properly

    # Step 2: Collect ALL custom label structures to preserve (EXCEPT CL2/INDEX2 and CL3/INDEX3)
    custom_label_structures = []
    custom_label_subdivisions = []

    if results:
        for row in results:
            criterion = row.ad_group_criterion
            lg = criterion.listing_group
            case_val = lg.case_value

            if (case_val and
                case_val._pb.WhichOneof("dimension") == "product_custom_attribute"):
                index_name = case_val.product_custom_attribute.index.name
                value = case_val.product_custom_attribute.value

                # Skip Custom Label 2 (INDEX2) and Custom Label 3 (INDEX3) - we're replacing them
                # INDEX2 is the old (incorrect) shop name targeting, INDEX3 is the new (correct) one
                if index_name == 'INDEX2' or index_name == 'INDEX3':
                    continue

                # Skip OTHERS cases (empty value)
                if not value or value == '':
                    continue

                # Collect SUBDIVISION nodes separately
                if lg.type_.name == 'SUBDIVISION':
                    custom_label_subdivisions.append({
                        'index': index_name,
                        'value': value,
                        'parent': lg.parent_ad_group_criterion if lg.parent_ad_group_criterion else None
                    })

                # Preserve all other custom label UNIT nodes (both negative and positive)
                if lg.type_.name == 'UNIT':
                    custom_label_structures.append({
                        'index': index_name,
                        'value': value,
                        'negative': criterion.negative,
                        'bid_micros': criterion.cpc_bid_micros
                    })

    if custom_label_subdivisions:
        print(f"      ℹ️ Found {len(custom_label_subdivisions)} existing subdivision(s):")
        for struct in custom_label_subdivisions:
            print(f"         - {struct['index']}: '{struct['value']}' (SUBDIVISION)")

    if custom_label_structures:
        print(f"      ℹ️ Preserving {len(custom_label_structures)} existing UNIT structure(s):")
        for struct in custom_label_structures:
            neg_str = "[NEGATIVE]" if struct['negative'] else "[POSITIVE]"
            print(f"         - {struct['index']}: '{struct['value']}' {neg_str}")

    # Step 3: Remove old tree
    safe_remove_entire_listing_tree(client, customer_id, str(ad_group_id))
    # No sleep needed - API operations are synchronous

    agc_service = client.get_service("AdGroupCriterionService")

    # Step 4: Rebuild tree hierarchically with preserved structures + CL3 exclusion
    # Use SUBDIVISIONS to determine hierarchy, not UNIT nodes

    # Group subdivisions by INDEX (dimension type)
    cl0_subdivisions = [s for s in custom_label_subdivisions if s['index'] == 'INDEX0']
    cl1_subdivisions = [s for s in custom_label_subdivisions if s['index'] == 'INDEX1']

    # Group UNIT structures by INDEX
    cl0_units = [s for s in custom_label_structures if s['index'] == 'INDEX0']
    cl1_units = [s for s in custom_label_structures if s['index'] == 'INDEX1']

    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # Determine hierarchy based on SUBDIVISIONS (not units)
    current_parent_tmp = root_tmp
    deepest_subdivision_tmp = root_tmp
    result_index_map = [0]  # Track which result index is which subdivision

    # If CL0 or CL1 subdivisions exist, rebuild them
    if cl0_subdivisions:
        # Build CL0 level
        cl0_subdiv = cl0_subdivisions[0]

        # Create CL0 subdivision
        dim_cl0 = client.get_type("ListingDimensionInfo")
        dim_cl0.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
        dim_cl0.product_custom_attribute.value = cl0_subdiv['value']

        cl0_subdivision_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=current_parent_tmp,
            listing_dimension_info=dim_cl0
        )
        cl0_subdivision_tmp = cl0_subdivision_op.create.resource_name
        ops1.append(cl0_subdivision_op)
        result_index_map.append(len(ops1) - 1)  # Track CL0 subdivision index

        # Add CL0 OTHERS (negative)
        dim_cl0_others = client.get_type("ListingDimensionInfo")
        dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=current_parent_tmp,
                listing_dimension_info=dim_cl0_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

        current_parent_tmp = cl0_subdivision_tmp
        deepest_subdivision_tmp = cl0_subdivision_tmp

    if cl1_subdivisions:
        # Build CL1 level under current parent
        cl1_subdiv = cl1_subdivisions[0]

        # Create CL1 subdivision
        dim_cl1 = client.get_type("ListingDimensionInfo")
        dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        dim_cl1.product_custom_attribute.value = cl1_subdiv['value']

        cl1_subdivision_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=current_parent_tmp,
            listing_dimension_info=dim_cl1
        )
        cl1_subdivision_tmp = cl1_subdivision_op.create.resource_name
        ops1.append(cl1_subdivision_op)
        result_index_map.append(len(ops1) - 1)  # Track CL1 subdivision index

        # Add CL1 OTHERS (negative)
        dim_cl1_others = client.get_type("ListingDimensionInfo")
        dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=current_parent_tmp,
                listing_dimension_info=dim_cl1_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

        current_parent_tmp = cl1_subdivision_tmp
        deepest_subdivision_tmp = cl1_subdivision_tmp

    # If there are CL0 units under the deepest subdivision, we need to convert them to subdivisions
    # and nest CL3 under them (following pattern from rebuild_tree_with_label_and_item_ids)
    if cl0_units:
        # For each CL0 unit, create as subdivision and add CL3 under it
        for unit in cl0_units:
            # Create CL0 subdivision (instead of unit)
            dim_cl0_subdiv = client.get_type("ListingDimensionInfo")
            dim_cl0_subdiv.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
            dim_cl0_subdiv.product_custom_attribute.value = unit['value']

            cl0_unit_subdivision_op = create_listing_group_subdivision(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_tmp,
                listing_dimension_info=dim_cl0_subdiv
            )
            cl0_unit_subdivision_tmp = cl0_unit_subdivision_op.create.resource_name
            ops1.append(cl0_unit_subdivision_op)

            # Add CL3 OTHERS under this CL0 subdivision
            dim_cl3_others = client.get_type("ListingDimensionInfo")
            dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            ops1.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=str(ad_group_id),
                    parent_ad_group_criterion_resource_name=cl0_unit_subdivision_tmp,
                    listing_dimension_info=dim_cl3_others,
                    targeting_negative=False,
                    cpc_bid_micros=unit['bid_micros']  # Use the original bid from CL0 unit
                )
            )

        # Add CL0 OTHERS (negative) under deepest subdivision
        dim_cl0_others = client.get_type("ListingDimensionInfo")
        dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_tmp,
                listing_dimension_info=dim_cl0_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )
    else:
        # No CL0 units - just add CL3 directly under deepest subdivision
        dim_cl3_others = client.get_type("ListingDimensionInfo")
        dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_tmp,
                listing_dimension_info=dim_cl3_others,
                targeting_negative=False,
                cpc_bid_micros=default_bid_micros
            )
        )

    # Execute first mutate
    try:
        resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    except Exception as e:
        print(f"   ❌ Error rebuilding tree: {e}")
        raise  # Re-raise exception so calling code can handle it properly

    # No sleep needed - API operations are synchronous

    # MUTATE 2: Add shop exclusion under each CL0 subdivision (if they exist)
    ops2 = []

    if cl0_units:
        # We created CL0 subdivisions - need to find their actual resource names and add exclusion to each
        # Calculate the index of the first CL0 subdivision in results
        base_index = 1  # Start after ROOT
        if cl0_subdivisions:
            base_index += 2  # CL0 subdivision + CL0 OTHERS
        if cl1_subdivisions:
            base_index += 2  # CL1 subdivision + CL1 OTHERS

        # Each CL0 unit became: CL0 subdivision + CL3 OTHERS
        # So CL0 subdivisions are at: base_index, base_index+2, base_index+4, ...
        for i, unit in enumerate(cl0_units):
            cl0_subdivision_actual = resp1.results[base_index + (i * 2)].resource_name

            dim_shop = client.get_type("ListingDimensionInfo")
            dim_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            dim_shop.product_custom_attribute.value = shop_name
            ops2.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=str(ad_group_id),
                    parent_ad_group_criterion_resource_name=cl0_subdivision_actual,
                    listing_dimension_info=dim_shop,
                    targeting_negative=True,
                    cpc_bid_micros=None
                )
            )
    else:
        # No CL0 units - add exclusion under the deepest subdivision (CL1 or ROOT)
        if cl1_subdivisions:
            if cl0_subdivisions:
                deepest_subdivision_actual = resp1.results[3].resource_name  # ROOT, CL0 subdivision, CL0 OTHERS, CL1 subdivision
            else:
                deepest_subdivision_actual = resp1.results[1].resource_name  # ROOT, CL1 subdivision
        elif cl0_subdivisions:
            deepest_subdivision_actual = resp1.results[1].resource_name  # ROOT, CL0 subdivision
        else:
            deepest_subdivision_actual = resp1.results[0].resource_name  # ROOT

        dim_shop = client.get_type("ListingDimensionInfo")
        dim_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
        dim_shop.product_custom_attribute.value = shop_name
        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_actual,
                listing_dimension_info=dim_shop,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    except Exception as e:
        print(f"   ❌ Error adding shop exclusion: {e}")
        raise  # Re-raise exception so calling code can handle it properly

    preserved_count = len(custom_label_structures)
    if preserved_count > 0:
        print(f"   ✅ Tree rebuilt: EXCLUDING shop '{shop_name}', preserved {preserved_count} existing structure(s)")
    else:
        print(f"   ✅ Tree rebuilt: EXCLUDING shop '{shop_name}', showing all others.")


def rebuild_tree_with_shop_exclusions(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: int,
    shop_names: list,
    required_cl0_value: str = None,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Rebuild listing tree with CL3 shop exclusions while preserving item ID exclusions.
    Validates and enforces CL0 and CL1 targeting based on Excel data and ad group name.

    Tree structure (with item IDs):
    ROOT (subdivision)
    ├─ CL0 = diepste_cat_id (subdivision) - from Excel column D
    │  ├─ CL1 = custom_label_1 (subdivision) - from ad group name suffix
    │  │  ├─ CL3 = shop1 (unit, negative) - exclude shop 1
    │  │  ├─ CL3 = shop2 (unit, negative) - exclude shop 2
    │  │  └─ CL3 OTHERS (subdivision) - for all other shops:
    │  │     ├─ ITEM_ID = xxx (unit, negative) - preserved exclusions
    │  │     ├─ ITEM_ID = yyy (unit, negative) - preserved exclusions
    │  │     └─ ITEM_ID OTHERS (unit, positive with bid)
    │  └─ CL1 OTHERS (unit, negative)
    └─ CL0 OTHERS (unit, negative)

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_names: List of shop names to exclude (CL3 values)
        required_cl0_value: Required CL0 value from Excel (diepste_cat_id)
        default_bid_micros: Bid amount in micros
    """
    print(f"   Rebuilding tree to EXCLUDE {len(shop_names)} shop(s): {', '.join(shop_names)}")

    # Step 1: Get ad group name to check for CL1 suffix requirement
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Query ad group name
    ag_name_query = f"""
        SELECT ad_group.name
        FROM ad_group
        WHERE ad_group.id = {ad_group_id}
    """

    try:
        ag_results = list(ga_service.search(customer_id=customer_id, query=ag_name_query))
        ad_group_name = ag_results[0].ad_group.name if ag_results else None
    except Exception as e:
        print(f"   ⚠️  Warning: Could not read ad group name: {e}")
        ad_group_name = None

    # Check if ad group name ends with _a, _b, or _c
    required_cl1 = None
    if ad_group_name:
        for suffix in ['_a', '_b', '_c']:
            if ad_group_name.endswith(suffix):
                required_cl1 = suffix[1:]  # Remove underscore: "_a" → "a"
                print(f"   📌 Ad group name ends with '{suffix}' → CL1 must be '{required_cl1}'")
                break

    # Step 2: Read existing tree to find CL0, CL1, and item ID exclusions
    query = f"""
        SELECT
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.case_value.product_item_id.value,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.cpc_bid_micros,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"   ❌ Error reading existing tree: {e}")
        raise

    # Extract CL0, CL1, item IDs, existing shop exclusions, and bid from existing tree
    cl0_value = None
    cl1_value = None
    existing_bid = default_bid_micros
    item_id_exclusions = []  # List of item IDs to preserve
    existing_shop_exclusions = []  # List of existing CL3 shop exclusions to preserve

    for row in results:
        case_value = row.ad_group_criterion.listing_group.case_value

        # Check for item ID
        if case_value.product_item_id.value:
            # Only preserve NEGATIVE item IDs (exclusions)
            if row.ad_group_criterion.negative:
                item_id_exclusions.append(case_value.product_item_id.value)

        # Check for custom attributes (CL0-CL4)
        if case_value.product_custom_attribute:
            index = case_value.product_custom_attribute.index.name
            value = case_value.product_custom_attribute.value

            # Get CL0 and CL1 from any node (subdivision or unit)
            if index == 'INDEX0' and value:
                cl0_value = value
            elif index == 'INDEX1' and value:
                cl1_value = value
            # Capture existing CL3 shop exclusions (NEGATIVE units with value, not OTHERS)
            elif index == 'INDEX3' and value:
                if (row.ad_group_criterion.listing_group.type.name == 'UNIT' and
                    row.ad_group_criterion.negative):
                    existing_shop_exclusions.append(value)

            # Capture existing bid from positive units only
            if (row.ad_group_criterion.listing_group.type.name == 'UNIT' and
                not row.ad_group_criterion.negative and
                row.ad_group_criterion.cpc_bid_micros):
                existing_bid = row.ad_group_criterion.cpc_bid_micros

    # Override CL0 if required value is specified from Excel
    if required_cl0_value:
        if cl0_value and cl0_value != required_cl0_value:
            print(f"   ⚠️  Overriding existing CL0='{cl0_value}' with required CL0='{required_cl0_value}' (from Excel diepste_cat_id)")
        cl0_value = required_cl0_value

    # Override CL1 if ad group name requires specific value
    if required_cl1:
        if cl1_value and cl1_value != required_cl1:
            print(f"   ⚠️  Overriding existing CL1='{cl1_value}' with required CL1='{required_cl1}' (from ad group name)")
        cl1_value = required_cl1

    # Validate we have required values
    if not cl0_value:
        if required_cl0_value:
            cl0_value = required_cl0_value
        else:
            raise Exception(f"Could not find CL0 value in existing tree and Excel doesn't specify one")
    if not cl1_value:
        raise Exception(f"Could not find CL1 value in existing tree and ad group name doesn't specify one")

    # Log what we found
    print(f"   Found existing structure: CL0={cl0_value}, CL1={cl1_value}, bid={existing_bid/10000:.2f}€")
    if existing_shop_exclusions:
        print(f"   Found {len(existing_shop_exclusions)} existing shop exclusion(s): {', '.join(existing_shop_exclusions)}")
    if item_id_exclusions:
        print(f"   Found {len(item_id_exclusions)} item ID exclusion(s)")

    # Merge new shop exclusions with existing ones (preserve all existing)
    # IMPORTANT: Use lowercase for comparison to avoid duplicates due to case differences
    existing_lower = {shop.lower(): shop for shop in existing_shop_exclusions}  # Map lowercase to original
    all_shop_exclusions = set(existing_shop_exclusions)  # Start with existing (preserve original case)
    new_shops_added = []

    for shop in shop_names:
        shop_lower = shop.lower()
        if shop_lower not in existing_lower:
            all_shop_exclusions.add(shop)
            existing_lower[shop_lower] = shop  # Track this one too
            new_shops_added.append(shop)

    if new_shops_added:
        print(f"   Adding {len(new_shops_added)} new shop exclusion(s): {', '.join(new_shops_added)}")
    else:
        print(f"   No new shop exclusions to add (all {len(shop_names)} already exist)")

    # Convert back to sorted list for consistent ordering (case-insensitive sort)
    shop_names = sorted(all_shop_exclusions, key=str.lower)
    print(f"   Total shop exclusions after merge: {len(shop_names)}")

    # Step 3: Remove entire tree
    safe_remove_entire_listing_tree(client, customer_id, str(ad_group_id))
    print(f"   Removed existing tree")

    # Step 4: Rebuild tree with shop exclusions and preserved item IDs
    has_item_ids = len(item_id_exclusions) > 0

    # Rebuild tree with multiple shop exclusions
    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create ROOT + CL0 subdivision + CL0 OTHERS (satisfies CL0) + ROOT OTHERS
    ops1 = []

    # ROOT subdivision
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # CL0 subdivision (under ROOT)
    dim_cl0 = client.get_type("ListingDimensionInfo")
    dim_cl0.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
    dim_cl0.product_custom_attribute.value = str(cl0_value)

    cl0_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl0
    )
    cl0_subdivision_tmp = cl0_subdivision_op.create.resource_name
    ops1.append(cl0_subdivision_op)

    # CL1 OTHERS (negative - under CL0) - This satisfies CL0 subdivision requirement
    dim_cl1_others_temp = client.get_type("ListingDimensionInfo")
    dim_cl1_others_temp.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=cl0_subdivision_tmp,  # Under CL0!
            listing_dimension_info=dim_cl1_others_temp,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # CL0 OTHERS (negative - under ROOT) - This satisfies ROOT subdivision requirement
    dim_cl0_others = client.get_type("ListingDimensionInfo")
    dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=root_tmp,  # Under ROOT
            listing_dimension_info=dim_cl0_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    try:
        response1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
        cl0_actual = response1.results[1].resource_name
    except Exception as e:
        raise Exception(f"Error creating ROOT and CL0: {e}")

    # MUTATE 2: Create CL1 subdivision + CL3 OTHERS (subdivision if item IDs, else unit)
    ops2 = []

    # CL1 subdivision (specific value, e.g., "b")
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    dim_cl1.product_custom_attribute.value = str(cl1_value)

    cl1_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=cl0_actual,
        listing_dimension_info=dim_cl1
    )
    cl1_subdivision_tmp = cl1_subdivision_op.create.resource_name
    ops2.append(cl1_subdivision_op)

    # CL3 OTHERS - subdivision if item IDs exist, else unit
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3

    if has_item_ids:
        # Create as SUBDIVISION to hold item ID exclusions underneath
        cl3_others_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=cl1_subdivision_tmp,
            listing_dimension_info=dim_cl3_others
        )
        cl3_others_tmp = cl3_others_op.create.resource_name
        ops2.append(cl3_others_op)

        # Add ITEM_ID OTHERS under CL3 OTHERS to satisfy subdivision requirement
        dim_item_others = client.get_type("ListingDimensionInfo")
        dim_item_others.product_item_id = client.get_type("ProductItemIdInfo")
        # Don't set value - this makes it OTHERS
        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=cl3_others_tmp,
                listing_dimension_info=dim_item_others,
                targeting_negative=False,  # Positive
                cpc_bid_micros=existing_bid
            )
        )
    else:
        # Create as UNIT with bid (no item IDs to preserve)
        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=cl1_subdivision_tmp,
                listing_dimension_info=dim_cl3_others,
                targeting_negative=False,  # Positive
                cpc_bid_micros=existing_bid
            )
        )

    try:
        response2 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
        cl1_actual = response2.results[0].resource_name
        if has_item_ids:
            cl3_others_actual = response2.results[1].resource_name  # Get actual CL3 OTHERS resource name
    except Exception as e:
        raise Exception(f"Error creating CL1 and CL3 OTHERS: {e}")

    # MUTATE 3: Add individual shop exclusions (CL3 OTHERS already exists from MUTATE 2)
    ops3 = []

    # Add each shop as a negative CL3 unit
    for shop in shop_names:
        dim_cl3_shop = client.get_type("ListingDimensionInfo")
        dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
        dim_cl3_shop.product_custom_attribute.value = str(shop)

        ops3.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=cl1_actual,
                listing_dimension_info=dim_cl3_shop,
                targeting_negative=True,  # NEGATIVE = exclude this shop
                cpc_bid_micros=None
            )
        )

    # Execute shop exclusions
    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops3)
    except Exception as e:
        raise Exception(f"Error adding shop exclusions: {e}")

    # MUTATE 4: Add item ID exclusions under CL3 OTHERS (if any exist)
    if has_item_ids:
        ops4 = []

        # Add each item ID as a negative unit under CL3 OTHERS
        for item_id in item_id_exclusions:
            dim_item_id = client.get_type("ListingDimensionInfo")
            dim_item_id.product_item_id = client.get_type("ProductItemIdInfo")
            dim_item_id.product_item_id.value = item_id

            ops4.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=str(ad_group_id),
                    parent_ad_group_criterion_resource_name=cl3_others_actual,
                    listing_dimension_info=dim_item_id,
                    targeting_negative=True,  # NEGATIVE = exclude this item ID
                    cpc_bid_micros=None
                )
            )

        # Execute item ID exclusions
        try:
            agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops4)
            print(f"   ✅ Tree rebuilt with {len(shop_names)} shop exclusion(s) and {len(item_id_exclusions)} item ID exclusion(s) preserved")
        except Exception as e:
            raise Exception(f"Error adding item ID exclusions: {e}")
    else:
        print(f"   ✅ Tree rebuilt with {len(shop_names)} shop exclusion(s)")


def build_listing_tree_for_inclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    custom_label_1: str,
    maincat_id: str,
    shop_name: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Build listing tree for inclusion logic (NEW STRUCTURE):

    Tree structure:
    ROOT (subdivision)
    ├─ Custom Label 3 = shop_name (subdivision)
    │  ├─ Custom Label 3 OTHERS (unit, negative)
    │  └─ Custom Label 4 = maincat_id (subdivision)
    │     ├─ Custom Label 4 OTHERS (unit, negative)
    │     ├─ Custom Label 1 = custom_label_1 (unit, biddable, positive) ← Added in MUTATE 2
    │     └─ Custom Label 1 OTHERS (unit, negative) ← Created in MUTATE 1 with temp name
    └─ Custom Label 3 OTHERS (unit, negative)

    CRITICAL: Google Ads requires that when you create a SUBDIVISION, you must
    provide its OTHERS case in the SAME mutate operation using temporary resource names.

    MUTATE 1: Create root + CL3 subdivision + CL4 subdivision + all OTHERS cases
    MUTATE 2: Add positive custom_label_1 target under maincat subdivision

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        custom_label_1: Custom label 1 value (a/b/c)
        maincat_id: Main category ID to target (custom label 4)
        shop_name: Shop name to target (custom label 3)
        default_bid_micros: Default bid in micros
    """
    print(f"      Building tree: Shop={shop_name}, Maincat ID={maincat_id}, CL1={custom_label_1}")

    # Check if listing tree already exists - if so, skip to preserve exclusions
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    check_query = f"""
        SELECT ad_group_criterion.resource_name
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
        LIMIT 1
    """

    try:
        existing_tree = list(ga_service.search(customer_id=customer_id, query=check_query))
        if existing_tree:
            print(f"      ℹ️  Listing tree already exists - skipping to preserve exclusions")
            return
    except Exception:
        pass  # No existing tree, proceed to create

    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create root + CL3 subdivision + CL4 subdivision + all OTHERS cases
    # CRITICAL: When creating a subdivision, you MUST provide its OTHERS case in the SAME mutate
    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # 2. Custom Label 3 subdivision (Custom Label 3 = shop_name)
    dim_cl3 = client.get_type("ListingDimensionInfo")
    dim_cl3.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3  # INDEX3 = Custom Label 3
    dim_cl3.product_custom_attribute.value = str(shop_name)

    cl3_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl3
    )
    cl3_subdivision_tmp = cl3_subdivision_op.create.resource_name
    ops1.append(cl3_subdivision_op)

    # 3. Custom Label 3 OTHERS (negative - blocks other shops)
    # This is a child of ROOT and satisfies the OTHERS requirement for root
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,  # NEGATIVE
            cpc_bid_micros=None
        )
    )

    # 4. Maincat ID subdivision (Custom Label 4 = maincat_id)
    # This is a child of CL3 subdivision (using TEMP name)
    dim_maincat = client.get_type("ListingDimensionInfo")
    dim_maincat.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4  # INDEX4 = Custom Label 4
    dim_maincat.product_custom_attribute.value = str(maincat_id)

    maincat_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=cl3_subdivision_tmp,  # Under CL3, not ROOT!
        listing_dimension_info=dim_maincat
    )
    maincat_subdivision_tmp = maincat_subdivision_op.create.resource_name
    ops1.append(maincat_subdivision_op)

    # 5. Custom Label 4 OTHERS (negative - blocks other categories)
    # This is a child of CL3 subdivision and satisfies the OTHERS requirement for CL3
    dim_cl4_others = client.get_type("ListingDimensionInfo")
    dim_cl4_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_subdivision_tmp,  # Child of CL3
            listing_dimension_info=dim_cl4_others,
            targeting_negative=True,  # NEGATIVE
            cpc_bid_micros=None
        )
    )

    # 6. Custom Label 1 OTHERS (negative - blocks other CL1 values)
    # This is a child of maincat_id subdivision (using TEMP name) and satisfies its OTHERS requirement
    dim_cl1_others = client.get_type("ListingDimensionInfo")
    dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=maincat_subdivision_tmp,  # Using TEMP name!
            listing_dimension_info=dim_cl1_others,
            targeting_negative=True,  # NEGATIVE - block other CL1 values
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    maincat_subdivision_actual = resp1.results[3].resource_name  # Fourth result is maincat subdivision (0=root, 1=cl3, 2=cl3_others, 3=cl4)

    # Small delay to prevent concurrent modification errors
    time.sleep(0.5)

    # MUTATE 2: Under maincat_id, add the positive custom_label_1 target
    # Note: CL1 OTHERS was already created in MUTATE 1
    ops2 = []

    # Custom Label 1 (Custom Label 1 = custom_label_1) - POSITIVE target
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1  # INDEX1 = Custom Label 1
    dim_cl1.product_custom_attribute.value = str(custom_label_1)

    ops2.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=maincat_subdivision_actual,
            listing_dimension_info=dim_cl1,
            targeting_negative=False,  # POSITIVE - target this CL1 value
            cpc_bid_micros=10_000  # 1 cent = €0.01 = 10,000 micros
        )
    )

    # Execute second mutate
    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    print(f"      ✅ Tree created: Shop '{shop_name}' → Maincat '{maincat_id}' → CL1 '{custom_label_1}'")


def build_listing_tree_for_inclusion_v2(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str,
    maincat_ids: list,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Build listing tree for inclusion logic (V2 - NEW STRUCTURE):

    Tree structure:
    ROOT (subdivision)
    ├─ Custom Label 3 = shop_name (subdivision)
    │  ├─ Custom Label 4 = maincat_id_1 (unit, biddable, positive)
    │  ├─ Custom Label 4 = maincat_id_2 (unit, biddable, positive)
    │  ├─ ... (more maincat_ids)
    │  └─ Custom Label 4 OTHERS (unit, negative)
    └─ Custom Label 3 OTHERS (unit, negative)

    Key differences from v1:
    - No CL1 targeting (simpler structure)
    - Multiple maincat_ids per ad group (all as positive units)
    - shop_name = ad_group_name (same value)

    IMPORTANT: This function will NOT rebuild the tree if one already exists,
    to preserve any existing exclusions that may have been added.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to target (custom label 3) - same as ad_group_name
        maincat_ids: List of maincat IDs to target (custom label 4)
        default_bid_micros: Default bid in micros
    """
    print(f"      Building tree: Shop={shop_name}, Maincat IDs={maincat_ids}")

    # Check if listing tree already exists - if so, skip to preserve exclusions
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    check_query = f"""
        SELECT ad_group_criterion.resource_name
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
        LIMIT 1
    """

    try:
        existing_tree = list(ga_service.search(customer_id=customer_id, query=check_query))
        if existing_tree:
            print(f"      ℹ️  Listing tree already exists - skipping to preserve exclusions")
            return
    except Exception:
        pass  # No existing tree, proceed to create

    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create ROOT + CL3 subdivision + CL3 OTHERS + CL4 OTHERS
    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # 2. Custom Label 3 subdivision (CL3 = shop_name)
    dim_cl3 = client.get_type("ListingDimensionInfo")
    dim_cl3.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3.product_custom_attribute.value = str(shop_name)

    cl3_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl3
    )
    cl3_subdivision_tmp = cl3_subdivision_op.create.resource_name
    ops1.append(cl3_subdivision_op)

    # 3. Custom Label 3 OTHERS (negative - blocks other shops)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # 4. Custom Label 4 OTHERS (negative - blocks other categories)
    # Must be created in same mutate as CL3 subdivision
    dim_cl4_others = client.get_type("ListingDimensionInfo")
    dim_cl4_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_subdivision_tmp,
            listing_dimension_info=dim_cl4_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    cl3_subdivision_actual = resp1.results[1].resource_name  # Second result is CL3 subdivision

    # MUTATE 2: Add all maincat_ids as positive CL4 units
    ops2 = []

    for maincat_id in maincat_ids:
        dim_cl4 = client.get_type("ListingDimensionInfo")
        dim_cl4.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
        dim_cl4.product_custom_attribute.value = str(maincat_id)

        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl3_subdivision_actual,
                listing_dimension_info=dim_cl4,
                targeting_negative=False,  # POSITIVE - target this maincat
                cpc_bid_micros=10_000  # 1 cent = €0.01 = 10,000 micros
            )
        )

    # Execute second mutate
    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    print(f"      ✅ Tree created: Shop '{shop_name}' → {len(maincat_ids)} maincat(s)")


def build_listing_tree_for_uitbreiding(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str,
    maincat_id: str,
    custom_label_1: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Build listing tree for uitbreiding (extension) logic:

    Tree structure:
    ROOT (subdivision)
    └─ CL1 = custom_label_1 (subdivision)
       ├─ CL3 = shop_name (subdivision)
       │  ├─ CL4 = maincat_id (unit, biddable, positive)
       │  └─ CL4 OTHERS (unit, negative)
       └─ CL3 OTHERS (unit, negative)
    └─ CL1 OTHERS (unit, negative)

    This targets:
    - Custom Label 1 = a/b/c (variant)
    - Custom Label 3 = shop_name
    - Custom Label 4 = maincat_id (category)

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to target (custom label 3)
        maincat_id: Category ID to target (custom label 4)
        custom_label_1: Label value (a/b/c) for custom label 1
        default_bid_micros: Default bid in micros
    """
    print(f"      Building tree: CL1={custom_label_1}, Shop={shop_name}, Maincat={maincat_id}")

    # Check if listing tree already exists - if so, skip to preserve exclusions
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    check_query = f"""
        SELECT ad_group_criterion.resource_name
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
        LIMIT 1
    """

    try:
        existing_tree = list(ga_service.search(customer_id=customer_id, query=check_query))
        if existing_tree:
            print(f"      ℹ️  Listing tree already exists - skipping to preserve exclusions")
            return
    except Exception:
        pass  # No existing tree, proceed to create

    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create ROOT + CL1 subdivision + CL1 OTHERS
    # Also need to add CL3 OTHERS under CL1 subdivision (required for subdivision)
    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # 2. Custom Label 1 subdivision (CL1 = a/b/c)
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    dim_cl1.product_custom_attribute.value = str(custom_label_1)

    cl1_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl1
    )
    cl1_subdivision_tmp = cl1_subdivision_op.create.resource_name
    ops1.append(cl1_subdivision_op)

    # 3. Custom Label 3 OTHERS under CL1 subdivision (required for CL1 subdivision)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl1_subdivision_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # 4. Custom Label 1 OTHERS (negative - blocks other variants)
    dim_cl1_others = client.get_type("ListingDimensionInfo")
    dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl1_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    cl1_subdivision_actual = resp1.results[1].resource_name  # Second result is CL1 subdivision

    # Wait for API to process before next mutate
    time.sleep(2)

    # MUTATE 2: Create CL3 subdivision under CL1 + CL4 OTHERS under CL3
    ops2 = []

    # 5. Custom Label 3 subdivision (CL3 = shop_name)
    dim_cl3 = client.get_type("ListingDimensionInfo")
    dim_cl3.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3.product_custom_attribute.value = str(shop_name)

    cl3_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=cl1_subdivision_actual,
        listing_dimension_info=dim_cl3
    )
    cl3_subdivision_tmp = cl3_subdivision_op.create.resource_name
    ops2.append(cl3_subdivision_op)

    # 6. Custom Label 4 OTHERS (negative - blocks other categories)
    dim_cl4_others = client.get_type("ListingDimensionInfo")
    dim_cl4_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4

    ops2.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_subdivision_tmp,
            listing_dimension_info=dim_cl4_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # Execute second mutate
    resp2 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    cl3_subdivision_actual = resp2.results[0].resource_name  # First result is CL3 subdivision

    # Wait for API to process before next mutate
    time.sleep(2)

    # MUTATE 3: Add maincat_id as positive CL4 unit
    ops3 = []

    dim_cl4 = client.get_type("ListingDimensionInfo")
    dim_cl4.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
    dim_cl4.product_custom_attribute.value = str(maincat_id)

    ops3.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_subdivision_actual,
            listing_dimension_info=dim_cl4,
            targeting_negative=False,  # POSITIVE - target this maincat
            cpc_bid_micros=10_000  # 1 cent = €0.01 = 10,000 micros
        )
    )

    # Execute third mutate
    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops3)
    print(f"      ✅ Tree created: CL1='{custom_label_1}' → CL3='{shop_name}' → CL4='{maincat_id}'")


# ============================================================================
# EXCEL PROCESSING
# ============================================================================

def process_inclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None
):
    """
    Process the 'toevoegen' (inclusion) sheet - V2 (NEW STRUCTURE).

    Excel columns:
    A. shop_name - Used to build ad group name: PLA/{shop_name}_{cl1}
    B. Shop ID (not used)
    C. maincat - Used to build campaign name: PLA/{maincat} store_{cl1}
    D. maincat_id - Used as Custom Label 4 (multiple per ad group)
    E. custom label 1 - Used in both campaign and ad group names
    F. budget (daily budget in EUR)
    G. result (TRUE/FALSE) - updated by script
    H. error message

    Naming conventions:
    - Campaign name: PLA/{maincat} store_{cl1}
    - Ad group name: PLA/{shop_name}_{cl1}

    Groups rows by derived campaign name (maincat + cl1), then by shop_name.
    For each campaign:
    1. Create campaign with derived name (status: PAUSED)
       - Uses budget from column F
       - Applies bid strategy from MCC based on custom_label_1
    2. For each unique shop_name within the campaign (status: ENABLED):
       - Create ad group with derived name
       - Collect all maincat_ids for that ad group
       - Build listing tree with shop_name as CL3
       - All maincat_ids as positive CL4 units
    3. Update column G (result) with TRUE/FALSE per row

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING INCLUSION SHEET (V2): '{SHEET_INCLUSION}'")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[SHEET_INCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_INCLUSION}' not found in workbook")
        return

    # Load a separate workbook with data_only=True to read calculated values from formulas
    # This is needed because cells may contain VLOOKUP formulas instead of plain values
    data_workbook = None
    data_sheet = None
    if file_path:
        try:
            data_workbook = load_workbook(file_path, data_only=True)
            data_sheet = data_workbook[SHEET_INCLUSION]
            print("   (Using data_only mode to read formula results)")
        except Exception as e:
            print(f"   ⚠️  Could not load data_only workbook: {e}")
            print(f"   (Will read formulas as-is - make sure cells contain values, not formulas)")

    # Column indices for this sheet
    COL_SHOP_NAME = 0      # A: shop_name
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat
    COL_MAINCAT_ID = 3     # D: maincat_id
    COL_CL1 = 4            # E: custom label 1
    COL_BUDGET = 5         # F: budget
    COL_RESULT = 6         # G: result (TRUE/FALSE)
    COL_ERR = 7            # H: error message

    # Step 1: Read all rows and group by campaign (maincat + cl1), then by shop_name
    campaigns = defaultdict(lambda: {
        'maincat': None,
        'cl1': None,
        'ad_groups': defaultdict(lambda: {'maincat_ids': set(), 'shop_id': None, 'rows': []}),
        'budget': None,
        'rows': []
    })

    print("Step 1: Reading and grouping rows by campaign (maincat + cl1) and ad group (shop_name)...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column is empty (use original sheet for status check)
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        # Read values from data_only sheet if available (to get formula results)
        # Otherwise fall back to original sheet
        if data_sheet:
            shop_name = data_sheet.cell(row=idx, column=COL_SHOP_NAME + 1).value
            shop_id = data_sheet.cell(row=idx, column=COL_SHOP_ID + 1).value
            maincat = data_sheet.cell(row=idx, column=COL_MAINCAT + 1).value
            maincat_id = data_sheet.cell(row=idx, column=COL_MAINCAT_ID + 1).value
            custom_label_1 = data_sheet.cell(row=idx, column=COL_CL1 + 1).value
            budget = data_sheet.cell(row=idx, column=COL_BUDGET + 1).value
        else:
            shop_name = row[COL_SHOP_NAME].value
            shop_id = row[COL_SHOP_ID].value
            maincat = row[COL_MAINCAT].value
            maincat_id = row[COL_MAINCAT_ID].value
            custom_label_1 = row[COL_CL1].value
            budget = row[COL_BUDGET].value

        # Validate required fields
        if not shop_name or not maincat or not maincat_id or not custom_label_1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/maincat_id/cl1), skipping")
            sheet.cell(row=idx, column=COL_RESULT + 1).value = False
            sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        # Build campaign name from maincat and cl1
        campaign_name = f"PLA/{maincat} store_{custom_label_1}"

        # Store campaign-level data
        campaigns[campaign_name]['maincat'] = maincat
        campaigns[campaign_name]['cl1'] = custom_label_1
        campaigns[campaign_name]['budget'] = budget
        campaigns[campaign_name]['rows'].append({'idx': idx, 'row': row})

        # Store ad group data - collect all maincat_ids for this shop
        campaigns[campaign_name]['ad_groups'][shop_name]['maincat_ids'].add(maincat_id)
        campaigns[campaign_name]['ad_groups'][shop_name]['shop_id'] = shop_id
        campaigns[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    print(f"   Found {len(campaigns)} campaign(s) to process")
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns.values())
    print(f"   Total ad groups: {total_ad_groups}\n")

    # Step 2: Process each campaign
    total_campaigns = len(campaigns)
    successful_campaigns = 0

    for campaign_idx, (campaign_name, campaign_data) in enumerate(campaigns.items(), start=1):
        print(f"\n{'─'*70}")
        print(f"CAMPAIGN {campaign_idx}/{total_campaigns}: {campaign_name}")
        print(f"{'─'*70}")

        budget_value = campaign_data['budget']
        custom_label_1 = campaign_data['cl1']
        maincat = campaign_data['maincat']
        ad_groups = campaign_data['ad_groups']

        print(f"   Maincat: {maincat}")
        print(f"   Budget: {budget_value} EUR")
        print(f"   Custom Label 1: {custom_label_1}")
        print(f"   Ad Groups (shops): {len(ad_groups)}")

        try:
            # Campaign configuration
            merchant_center_account_id = 140784594  # Merchant Center ID
            budget_name = f"Budget_{campaign_name}"
            tracking_template = ""
            country = "NL"

            # Convert budget from EUR to micros
            try:
                budget_micros = int(float(budget_value) * 1_000_000) if budget_value else 10_000_000
            except (ValueError, TypeError):
                print(f"   ⚠️  Invalid budget value '{budget_value}', using default 10 EUR")
                budget_micros = 10_000_000

            # Get bid strategy based on custom label 1
            bid_strategy_resource_name = None
            if custom_label_1 and custom_label_1 in BID_STRATEGY_MAPPING:
                bid_strategy_name = BID_STRATEGY_MAPPING[custom_label_1]
                print(f"   Looking up bid strategy: {bid_strategy_name}")
                bid_strategy_resource_name = get_bid_strategy_by_name(
                    client=client,
                    customer_id=MCC_ACCOUNT_ID,
                    strategy_name=bid_strategy_name
                )

            # Get first ad group's shop info for campaign metadata
            first_ag_name = list(ad_groups.keys())[0]
            first_ag_data = ad_groups[first_ag_name]

            # Create campaign (status: PAUSED - set in add_standard_shopping_campaign)
            print(f"\n   Creating campaign: {campaign_name}")
            campaign_resource_name = add_standard_shopping_campaign(
                client=client,
                customer_id=customer_id,
                merchant_center_account_id=merchant_center_account_id,
                campaign_name=campaign_name,
                budget_name=budget_name,
                tracking_template=tracking_template,
                country=country,
                shopid=first_ag_data['shop_id'],
                shopname=first_ag_name,
                label=custom_label_1,
                budget=budget_micros,
                bidding_strategy_resource_name=bid_strategy_resource_name
            )

            if not campaign_resource_name:
                raise Exception("Failed to create/find campaign")

            print(f"   ✅ Campaign ready: {campaign_resource_name}")

            # Add negative keyword list to campaign
            if NEGATIVE_LIST_NAME:
                enable_negative_list_for_campaign(
                    client=client,
                    customer_id=customer_id,
                    campaign_resource_name=campaign_resource_name,
                    negative_list_name=NEGATIVE_LIST_NAME
                )

            # Wait after campaign setup before processing ad groups
            time.sleep(1.0)

            # Process each ad group (shop) within this campaign
            print(f"\n   Processing {len(ad_groups)} ad group(s)...")
            ad_groups_processed = []
            ad_group_errors = {}

            for ag_idx, (shop_name, ag_data) in enumerate(ad_groups.items(), start=1):
                # Build ad group name: PLA/{shop_name}_{cl1}
                ad_group_name = f"PLA/{shop_name}_{custom_label_1}"
                print(f"\n   ──── Ad Group {ag_idx}/{len(ad_groups)}: {ad_group_name} ────")
                print(f"      (Shop: {shop_name})")

                try:
                    maincat_ids = sorted(ag_data['maincat_ids'])
                    print(f"      Maincat IDs (CL4): {maincat_ids}")

                    # Create ad group (status: ENABLED - set in add_shopping_ad_group)
                    ad_group_resource_name, _ = add_shopping_ad_group(
                        client=client,
                        customer_id=customer_id,
                        campaign_resource_name=campaign_resource_name,
                        ad_group_name=ad_group_name,
                        campaign_name=campaign_name
                    )

                    if not ad_group_resource_name:
                        raise Exception(f"Failed to create/find ad group")

                    print(f"      ✅ Ad group ready: {ad_group_resource_name}")

                    # Wait after ad group creation before building tree
                    time.sleep(1.0)

                    # Extract ad group ID
                    ad_group_id = ad_group_resource_name.split('/')[-1]

                    # For CL3 targeting, split shop_name at | and use first part
                    # e.g. "Hbm-machines.com|NL" becomes "Hbm-machines.com"
                    shop_name_for_targeting = shop_name.split('|')[0] if '|' in shop_name else shop_name
                    if shop_name_for_targeting != shop_name:
                        print(f"      CL3 targeting: '{shop_name_for_targeting}' (split from '{shop_name}')")

                    # Build listing tree with V2 function
                    build_listing_tree_for_inclusion_v2(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        shop_name=shop_name_for_targeting,  # Use split shop_name for CL3
                        maincat_ids=maincat_ids
                    )

                    # Wait after tree creation before creating ad
                    time.sleep(2.0)

                    # Create shopping product ad
                    print(f"      Creating shopping product ad...")
                    add_shopping_product_ad(
                        client=client,
                        customer_id=customer_id,
                        ad_group_resource_name=ad_group_resource_name
                    )

                    ad_groups_processed.append(shop_name)
                    print(f"      ✅ Ad group completed: {ad_group_name}")

                    # Wait between ad groups to prevent concurrent modification
                    time.sleep(1.0)

                except Exception as e:
                    error_msg = str(e)
                    print(f"      ❌ Failed: {error_msg}")
                    ad_group_errors[shop_name] = error_msg

            # Mark rows as successful/failed
            for shop_name, ag_data in ad_groups.items():
                for row_info in ag_data['rows']:
                    row_num = row_info['idx']
                    if shop_name in ad_groups_processed:
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = ""
                    else:
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                        error_msg = ad_group_errors.get(shop_name, "Failed to process ad group")
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = error_msg[:100]

            if len(ad_groups_processed) > 0:
                successful_campaigns += 1
                print(f"\n   ✅ CAMPAIGN COMPLETED: {len(ad_groups_processed)}/{len(ad_groups)} ad groups processed")

        except Exception as e:
            error_msg = str(e)
            print(f"\n   ❌ CAMPAIGN FAILED: {error_msg}")
            # Mark all rows for this campaign as failed
            for row_info in campaign_data['rows']:
                row_num = row_info['idx']
                sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                sheet.cell(row=row_num, column=COL_ERR + 1).value = f"Campaign failed: {error_msg[:80]}"

        # Save periodically
        if file_path and campaign_idx % 5 == 0:
            print(f"\n   💾 Saving progress...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"   ⚠️  Error saving: {save_error}")

        # Wait between campaigns to prevent concurrent modification
        time.sleep(2.0)

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    # Close data_only workbook if it was opened
    if data_workbook:
        data_workbook.close()

    print(f"\n{'='*70}")
    print(f"INCLUSION SHEET (V2) SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaigns: {total_campaigns}")
    print(f"✅ Successful: {successful_campaigns}")
    print(f"❌ Failed: {total_campaigns - successful_campaigns}")
    print(f"{'='*70}\n")


def pause_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_resource_name: str
) -> bool:
    """
    Pause an ad group by setting its status to PAUSED.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_resource_name: Resource name of the ad group to pause

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        from google.protobuf import field_mask_pb2

        ad_group_service = client.get_service("AdGroupService")
        ad_group_operation = client.get_type("AdGroupOperation")

        ad_group = ad_group_operation.update
        ad_group.resource_name = ad_group_resource_name
        ad_group.status = client.enums.AdGroupStatusEnum.PAUSED

        # Set the field mask to only update the status field
        ad_group_operation.update_mask.CopyFrom(
            field_mask_pb2.FieldMask(paths=["status"])
        )

        ad_group_service.mutate_ad_groups(
            customer_id=customer_id,
            operations=[ad_group_operation]
        )

        return True
    except GoogleAdsException as ex:
        print(f"      ❌ Google Ads API error: {ex.error.code().name}")
        for error in ex.failure.errors:
            print(f"         {error.message}")
        return False
    except Exception as e:
        print(f"      ❌ Error pausing ad group: {str(e)}")
        return False


def enable_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_resource_name: str
) -> bool:
    """
    Enable an ad group by setting its status to ENABLED.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_resource_name: Resource name of the ad group to enable

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        from google.protobuf import field_mask_pb2

        ad_group_service = client.get_service("AdGroupService")
        ad_group_operation = client.get_type("AdGroupOperation")

        ad_group = ad_group_operation.update
        ad_group.resource_name = ad_group_resource_name
        ad_group.status = client.enums.AdGroupStatusEnum.ENABLED

        # Set the field mask to only update the status field
        ad_group_operation.update_mask.CopyFrom(
            field_mask_pb2.FieldMask(paths=["status"])
        )

        ad_group_service.mutate_ad_groups(
            customer_id=customer_id,
            operations=[ad_group_operation]
        )

        return True
    except GoogleAdsException as ex:
        print(f"      ❌ Google Ads API error: {ex.error.code().name}")
        for error in ex.failure.errors:
            print(f"         {error.message}")
        return False
    except Exception as e:
        print(f"      ❌ Error enabling ad group: {str(e)}")
        return False


def remove_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_resource_name: str
) -> bool:
    """
    Remove (delete) an ad group using the remove operation.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_resource_name: Resource name of the ad group to remove

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        ad_group_service = client.get_service("AdGroupService")
        ad_group_operation = client.get_type("AdGroupOperation")

        # Use the remove operation instead of update with REMOVED status
        ad_group_operation.remove = ad_group_resource_name

        ad_group_service.mutate_ad_groups(
            customer_id=customer_id,
            operations=[ad_group_operation]
        )

        return True
    except GoogleAdsException as ex:
        print(f"      ❌ Google Ads API error: {ex.error.code().name}")
        for error in ex.failure.errors:
            print(f"         {error.message}")
        return False
    except Exception as e:
        print(f"      ❌ Error removing ad group: {str(e)}")
        return False


def find_ad_group_in_campaign(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_name: str,
    ad_group_name: str
) -> Optional[Dict[str, Any]]:
    """
    Find an ad group by campaign name and ad group name.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_name: Name of the campaign
        ad_group_name: Name of the ad group

    Returns:
        dict with ad_group info or None if not found
    """
    google_ads_service = client.get_service("GoogleAdsService")

    # Escape special characters for GAQL (only single quotes need escaping)
    def escape_gaql_string(s):
        s = s.replace("'", "\\'")
        return s

    escaped_campaign_name = escape_gaql_string(campaign_name)
    escaped_ad_group_name = escape_gaql_string(ad_group_name)

    query = f"""
        SELECT
            ad_group.id,
            ad_group.resource_name,
            ad_group.name,
            ad_group.status,
            campaign.id,
            campaign.name,
            campaign.resource_name,
            campaign.status
        FROM ad_group
        WHERE campaign.name = '{escaped_campaign_name}'
        AND ad_group.name = '{escaped_ad_group_name}'
        AND ad_group.status IN ('ENABLED', 'PAUSED')
        AND campaign.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = google_ads_service.search(customer_id=customer_id, query=query)
        for row in response:
            return {
                'ad_group_id': row.ad_group.id,
                'ad_group_resource_name': row.ad_group.resource_name,
                'ad_group_name': row.ad_group.name,
                'ad_group_status': row.ad_group.status.name,
                'campaign_id': row.campaign.id,
                'campaign_name': row.campaign.name,
                'campaign_resource_name': row.campaign.resource_name
            }
    except GoogleAdsException as ex:
        print(f"      ❌ Error searching for ad group: {ex}")

    return None


def process_reverse_inclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None
):
    """
    Process the 'toevoegen' sheet - removes (deletes) ad groups.

    This is the reverse of process_inclusion_sheet_v2. Instead of creating
    ad groups, it finds existing ad groups and removes them.

    Excel columns:
    A. shop_name - Used to build ad group name: PLA/{shop_name}_{cl1}
    B. Shop ID (not used)
    C. maincat - Used to build campaign name: PLA/{maincat} store_{cl1}
    D. maincat_id (not used)
    E. custom label 1 - Used in both campaign and ad group names
    F. budget (ignored)
    G. result (TRUE/FALSE) - updated by script
    H. Error message

    Naming conventions:
    - Campaign name: PLA/{maincat} store_{cl1}
    - Ad group name: PLA/{shop_name}_{cl1}

    Groups rows by derived campaign name (from maincat + cl1), then removes each
    unique ad group (derived from shop_name + cl1) within that campaign.

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING REVERSE INCLUSION SHEET (V2): '{SHEET_REVERSE_INCLUSION}'")
    print(f"(REMOVING AD GROUPS)")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[SHEET_REVERSE_INCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_REVERSE_INCLUSION}' not found in workbook")
        return

    # Column indices for toevoegen sheet
    COL_SHOP_NAME = 0      # A: Shop name (= ad group name)
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat (category name)
    COL_MAINCAT_ID = 3     # D: maincat_id (not used)
    COL_CL1 = 4            # E: custom label 1
    COL_BUDGET = 5         # F: budget (ignored)
    COL_RESULT = 6         # G: result (TRUE/FALSE)
    COL_ERR = 7            # H: Error message

    # Step 1: Read all rows and group by campaign (derived from maincat + cl1)
    # Each shop_name is an ad group within that campaign
    campaigns_to_process = defaultdict(lambda: {
        'maincat': None,
        'cl1': None,
        'ad_groups': defaultdict(lambda: {'rows': []})  # shop_name -> rows
    })

    print("Step 1: Reading and grouping rows by campaign (maincat + cl1) and ad group (shop_name)...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column is empty
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value  # This is the ad group name
        maincat = row[COL_MAINCAT].value
        cl1 = row[COL_CL1].value

        # Validate required fields
        if not shop_name or not maincat or not cl1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/cl1), skipping")
            sheet.cell(row=idx, column=COL_RESULT + 1).value = False
            sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        # Build campaign name from maincat and cl1
        campaign_name = f"PLA/{maincat} store_{cl1}"

        # Group by campaign, then by ad group (shop_name)
        campaigns_to_process[campaign_name]['maincat'] = maincat
        campaigns_to_process[campaign_name]['cl1'] = cl1
        campaigns_to_process[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    total_campaigns = len(campaigns_to_process)
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns_to_process.values())
    print(f"   Found {total_campaigns} campaign(s) with {total_ad_groups} unique ad group(s) to remove")

    # Step 2: Process each campaign and its ad groups
    successful_removals = 0
    failed_removals = 0
    processed_ag_count = 0

    for camp_idx, (campaign_name, campaign_data) in enumerate(campaigns_to_process.items(), start=1):
        print(f"\n{'─'*70}")
        print(f"CAMPAIGN {camp_idx}/{total_campaigns}: {campaign_name}")
        print(f"{'─'*70}")
        print(f"   Maincat: {campaign_data['maincat']}")
        print(f"   Custom Label 1: {campaign_data['cl1']}")
        print(f"   Ad Groups to remove: {len(campaign_data['ad_groups'])}")

        # Process each ad group (shop_name) in this campaign
        for shop_name, ag_data in campaign_data['ad_groups'].items():
            processed_ag_count += 1

            # Build ad group name: PLA/{shop_name}_{cl1}
            ad_group_name = f"PLA/{shop_name}_{campaign_data['cl1']}"
            print(f"\n   ──── Ad Group: {ad_group_name} ────")
            print(f"      (Shop: {shop_name})")

            try:
                # Find the ad group
                print(f"      Searching for ad group...")
                ad_group_info = find_ad_group_in_campaign(
                    client=client,
                    customer_id=customer_id,
                    campaign_name=campaign_name,
                    ad_group_name=ad_group_name
                )

                if not ad_group_info:
                    raise Exception(f"Ad group not found in campaign")

                print(f"      ✅ Found ad group (ID: {ad_group_info['ad_group_id']})")
                print(f"         Current status: {ad_group_info['ad_group_status']}")

                # Check if already removed
                if ad_group_info['ad_group_status'] == 'REMOVED':
                    print(f"      ℹ️  Ad group is already REMOVED")
                    # Mark as successful anyway
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = "Already removed"
                    successful_removals += 1
                    continue

                # Remove the ad group
                print(f"      Removing ad group...")
                success = remove_ad_group(
                    client=client,
                    customer_id=customer_id,
                    ad_group_resource_name=ad_group_info['ad_group_resource_name']
                )

                if success:
                    print(f"      ✅ Ad group removed successfully")
                    successful_removals += 1
                    # Mark all rows for this ad group as successful
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = ""
                else:
                    raise Exception("Failed to remove ad group")

                time.sleep(0.3)  # Small delay between API calls

            except Exception as e:
                error_msg = str(e)
                print(f"      ❌ Failed: {error_msg}")
                failed_removals += 1
                # Mark all rows for this ad group as failed
                for row_info in ag_data['rows']:
                    row_num = row_info['idx']
                    sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_num, column=COL_ERR + 1).value = error_msg[:100]

            # Save periodically
            if file_path and processed_ag_count % 10 == 0:
                print(f"\n   💾 Saving progress...")
                try:
                    workbook.save(file_path)
                except Exception as save_error:
                    print(f"   ⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"REVERSE INCLUSION SHEET (V2) SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaigns: {total_campaigns}")
    print(f"Total ad groups: {total_ad_groups}")
    print(f"✅ Removed: {successful_removals}")
    print(f"❌ Failed: {failed_removals}")
    print(f"{'='*70}\n")


def process_enable_inclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    sheet_name: str = "adgroups_heractiveren"
):
    """
    Process a sheet to ENABLE ad groups (reverse of pause).

    This enables previously paused ad groups, bringing them back online.

    Excel columns (adgroups_heractiveren):
    A. shop_name - Used to build ad group name: PLA/{shop_name}_{cl1}
    B. Shop ID (not used)
    C. maincat - Used to build campaign name: PLA/{maincat} store_{cl1}
    D. maincat_id (not used)
    E. custom label 1 - Used in both campaign and ad group names
    F. budget (ignored)
    G. result (TRUE/FALSE) - updated by script
    H. Error message

    Naming conventions:
    - Campaign name: PLA/{maincat} store_{cl1}
    - Ad group name: PLA/{shop_name}_{cl1}

    Groups rows by derived campaign name (from maincat + cl1), then enables each
    unique ad group (derived from shop_name + cl1) within that campaign.

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        sheet_name: Name of sheet to process (default: 'hervatten')
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING ENABLE INCLUSION SHEET (V2): '{sheet_name}'")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"❌ Sheet '{sheet_name}' not found in workbook")
        return

    # Column indices for this sheet
    COL_SHOP_NAME = 0      # A: Shop name (= ad group name)
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat (category name)
    COL_MAINCAT_ID = 3     # D: maincat_id (not used)
    COL_CL1 = 4            # E: custom label 1
    COL_BUDGET = 5         # F: budget (ignored)
    COL_RESULT = 6         # G: result (TRUE/FALSE)
    COL_ERR = 7            # H: Error message

    # Step 1: Read all rows and group by campaign (derived from maincat + cl1)
    # Each shop_name is an ad group within that campaign
    campaigns_to_process = defaultdict(lambda: {
        'maincat': None,
        'cl1': None,
        'ad_groups': defaultdict(lambda: {'rows': []})  # shop_name -> rows
    })

    print("Step 1: Reading and grouping rows by campaign (maincat + cl1) and ad group (shop_name)...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column is empty
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value  # This is the ad group name
        maincat = row[COL_MAINCAT].value
        cl1 = row[COL_CL1].value

        # Validate required fields
        if not shop_name or not maincat or not cl1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/cl1), skipping")
            sheet.cell(row=idx, column=COL_RESULT + 1).value = False
            sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        # Build campaign name from maincat and cl1
        campaign_name = f"PLA/{maincat} store_{cl1}"

        # Group by campaign, then by ad group (shop_name)
        campaigns_to_process[campaign_name]['maincat'] = maincat
        campaigns_to_process[campaign_name]['cl1'] = cl1
        campaigns_to_process[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    total_campaigns = len(campaigns_to_process)
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns_to_process.values())
    print(f"   Found {total_campaigns} campaign(s) with {total_ad_groups} unique ad group(s) to enable")

    # Step 2: Process each campaign and its ad groups
    successful_enables = 0
    failed_enables = 0
    processed_ag_count = 0

    for camp_idx, (campaign_name, campaign_data) in enumerate(campaigns_to_process.items(), start=1):
        print(f"\n{'─'*70}")
        print(f"CAMPAIGN {camp_idx}/{total_campaigns}: {campaign_name}")
        print(f"{'─'*70}")
        print(f"   Maincat: {campaign_data['maincat']}")
        print(f"   Custom Label 1: {campaign_data['cl1']}")
        print(f"   Ad Groups to enable: {len(campaign_data['ad_groups'])}")

        # Process each ad group (shop_name) in this campaign
        for shop_name, ag_data in campaign_data['ad_groups'].items():
            processed_ag_count += 1

            # Build ad group name: PLA/{shop_name}_{cl1}
            ad_group_name = f"PLA/{shop_name}_{campaign_data['cl1']}"
            print(f"\n   ──── Ad Group: {ad_group_name} ────")
            print(f"      (Shop: {shop_name})")

            try:
                # Find the ad group
                print(f"      Searching for ad group...")
                ad_group_info = find_ad_group_in_campaign(
                    client=client,
                    customer_id=customer_id,
                    campaign_name=campaign_name,
                    ad_group_name=ad_group_name
                )

                if not ad_group_info:
                    raise Exception(f"Ad group not found in campaign")

                print(f"      ✅ Found ad group (ID: {ad_group_info['ad_group_id']})")
                print(f"         Current status: {ad_group_info['ad_group_status']}")

                # Check if already enabled
                if ad_group_info['ad_group_status'] == 'ENABLED':
                    print(f"      ℹ️  Ad group is already ENABLED")
                    # Mark as successful anyway
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = "Already enabled"
                    successful_enables += 1
                    continue

                # Enable the ad group
                print(f"      Enabling ad group...")
                success = enable_ad_group(
                    client=client,
                    customer_id=customer_id,
                    ad_group_resource_name=ad_group_info['ad_group_resource_name']
                )

                if success:
                    print(f"      ✅ Ad group enabled successfully")
                    successful_enables += 1
                    # Mark all rows for this ad group as successful
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = ""
                else:
                    raise Exception("Failed to enable ad group")

                time.sleep(0.3)  # Small delay between API calls

            except Exception as e:
                error_msg = str(e)
                print(f"      ❌ Failed: {error_msg}")
                failed_enables += 1
                # Mark all rows for this ad group as failed
                for row_info in ag_data['rows']:
                    row_num = row_info['idx']
                    sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_num, column=COL_ERR + 1).value = error_msg[:100]

            # Save periodically
            if file_path and processed_ag_count % 10 == 0:
                print(f"\n   💾 Saving progress...")
                try:
                    workbook.save(file_path)
                except Exception as save_error:
                    print(f"   ⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"ENABLE INCLUSION SHEET (V2) SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaigns: {total_campaigns}")
    print(f"Total ad groups: {total_ad_groups}")
    print(f"✅ Enabled: {successful_enables}")
    print(f"❌ Failed: {failed_enables}")
    print(f"{'='*70}\n")


def process_inclusion_sheet_legacy(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None
):
    """
    Process the 'toevoegen' (inclusion) sheet - LEGACY VERSION.

    Excel columns (OLD):
    A. Shop name
    B. Shop ID
    C. maincat
    D. maincat_id
    E. custom label 1
    F. budget (daily budget in EUR)
    G. Status (TRUE/FALSE) - updated by script

    Groups rows by unique combination of (maincat, custom_label_1) ONLY.
    For each group:
    1. Create ONE campaign with name: PLA/{maincat} store_{custom_label_1}
       - Uses budget from column F (converted to micros)
       - Applies bid strategy from MCC based on custom_label_1
    2. Create MULTIPLE ad groups (one per unique shop_name in group)
       - Ad group names: PLA/{shop_name}_{custom_label_1}
    3. Build listing tree for EACH ad group:
       - Target maincat_id as custom label 4
       - Subdivide and target shop_name as custom label 3
       - Exclude everything else at both levels
       - Bid: 1 cent (10,000 micros)
    4. Update column G (status) with TRUE/FALSE per row based on shop success

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving progress)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING INCLUSION SHEET (LEGACY): '{SHEET_INCLUSION}'")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[SHEET_INCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_INCLUSION}' not found in workbook")
        return

    # Step 1: Read all rows and group by (maincat, custom_label_1) only
    groups = defaultdict(list)  # key: (maincat, custom_label_1), value: list of row data

    print("Step 1: Reading and grouping rows...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column (G) is empty - if so, this is where we start processing
        status_value = row[COL_LEGACY_STATUS].value

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_LEGACY_SHOP_NAME].value
        shop_id = row[COL_LEGACY_SHOP_ID].value
        maincat = row[COL_LEGACY_MAINCAT].value
        maincat_id = row[COL_LEGACY_MAINCAT_ID].value
        custom_label_1 = row[COL_LEGACY_CUSTOM_LABEL_1].value
        budget = row[COL_LEGACY_BUDGET].value

        # Validate required fields
        if not shop_name or not maincat or not maincat_id or not custom_label_1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/maincat_id/custom_label_1), skipping")
            row[COL_LEGACY_STATUS].value = False
            # Only write to error column if it exists
            if len(row) > COL_LEGACY_ERROR:
                row[COL_LEGACY_ERROR].value = "Missing required fields (shop_name/maincat/maincat_id/custom_label_1)"
            continue

        # Group by (maincat, custom_label_1) only - multiple shops per campaign
        group_key = (maincat, custom_label_1)

        # Store row data
        groups[group_key].append({
            'row_idx': idx,
            'row_obj': row,
            'shop_name': shop_name,
            'shop_id': shop_id,
            'maincat': maincat,
            'maincat_id': maincat_id,
            'custom_label_1': custom_label_1,
            'budget': budget
        })

    print(f"   Found {len(groups)} unique group(s) to process\n")

    # Step 2: Process each group
    total_groups = len(groups)
    successful_groups = 0

    for group_idx, (group_key, rows_in_group) in enumerate(groups.items(), start=1):
        # Delay between groups to prevent concurrent modification errors
        if group_idx > 1:
            time.sleep(2)

        maincat, custom_label_1 = group_key

        print(f"\n{'─'*70}")
        print(f"GROUP {group_idx}/{total_groups}: {maincat} | {custom_label_1}")
        print(f"   Rows in group: {len(rows_in_group)}")
        print(f"{'─'*70}")

        # Get metadata from first row (all rows in group share same maincat, maincat_id, budget)
        first_row = rows_in_group[0]
        maincat_id = first_row['maincat_id']
        budget_value = first_row['budget']

        # Get unique shops in this group
        unique_shops = {}  # shop_name -> shop_id mapping
        for row_data in rows_in_group:
            unique_shops[row_data['shop_name']] = row_data['shop_id']

        print(f"   Maincat ID: {maincat_id}")
        print(f"   Budget: {budget_value} EUR")
        print(f"   Unique shops in group: {len(unique_shops)}")

        try:
            # Build campaign name: PLA/{maincat} store_{custom_label_1}
            campaign_name = f"PLA/{maincat} store_{custom_label_1}"
            print(f"\n   Step 1: Checking for existing campaign or creating new: {campaign_name}")

            # Campaign configuration
            merchant_center_account_id = 140784594  # Merchant Center ID
            budget_name = f"Budget_{campaign_name}"
            tracking_template = ""  # Not needed
            country = "NL"  # Always Netherlands

            # Convert budget from EUR to micros (EUR * 1,000,000)
            # Default to 10 EUR if budget is missing or invalid
            try:
                budget_micros = int(float(budget_value) * 1_000_000) if budget_value else 10_000_000
            except (ValueError, TypeError):
                print(f"   ⚠️  Invalid budget value '{budget_value}', using default 10 EUR")
                budget_micros = 10_000_000

            # Get bid strategy based on custom label 1 (from MCC account)
            bid_strategy_resource_name = None
            if custom_label_1 in BID_STRATEGY_MAPPING:
                bid_strategy_name = BID_STRATEGY_MAPPING[custom_label_1]
                print(f"   Looking up bid strategy: {bid_strategy_name} (in MCC account)")
                bid_strategy_resource_name = get_bid_strategy_by_name(
                    client=client,
                    customer_id=MCC_ACCOUNT_ID,  # Search in MCC account
                    strategy_name=bid_strategy_name
                )

            # Use first shop's ID for campaign metadata
            first_shop_id = list(unique_shops.values())[0]
            first_shop_name = list(unique_shops.keys())[0]

            campaign_resource_name = add_standard_shopping_campaign(
                client=client,
                customer_id=customer_id,
                merchant_center_account_id=merchant_center_account_id,
                campaign_name=campaign_name,
                budget_name=budget_name,
                tracking_template=tracking_template,
                country=country,
                shopid=first_shop_id,
                shopname=first_shop_name,
                label=custom_label_1,
                budget=budget_micros,
                bidding_strategy_resource_name=bid_strategy_resource_name
            )

            if not campaign_resource_name:
                raise Exception("Failed to create/find campaign")

            print(f"   Campaign resource: {campaign_resource_name}")

            # Check/create multiple ad groups - one for each unique shop
            print(f"\n   Step 2: Processing ad groups for {len(unique_shops)} shop(s)...")
            shops_processed_successfully = []
            shop_errors = {}  # Track errors per shop

            for shop_idx, (shop_name, shop_id) in enumerate(unique_shops.items(), start=1):
                # Delay between shops to prevent concurrent modification errors
                if shop_idx > 1:
                    time.sleep(1.5)

                print(f"\n   ──── Shop {shop_idx}/{len(unique_shops)}: {shop_name} ────")

                try:
                    # Build ad group name: PLA/{shop_name}_{custom_label_1}
                    ad_group_name = f"PLA/{shop_name}_{custom_label_1}"
                    print(f"      Checking/creating ad group: {ad_group_name}")

                    ad_group_resource_name, _ = add_shopping_ad_group(
                        client=client,
                        customer_id=customer_id,
                        campaign_resource_name=campaign_resource_name,
                        ad_group_name=ad_group_name,
                        campaign_name=campaign_name
                    )

                    if not ad_group_resource_name:
                        raise Exception(f"Failed to create/find ad group for {shop_name}")

                    print(f"      ✅ Ad group ready: {ad_group_resource_name}")

                    # Extract ad group ID from resource name
                    ad_group_id = ad_group_resource_name.split('/')[-1]

                    # Build listing tree for this shop
                    print(f"      Building listing tree...")
                    build_listing_tree_for_inclusion(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        custom_label_1=custom_label_1,
                        maincat_id=maincat_id,
                        shop_name=shop_name,
                        default_bid_micros=DEFAULT_BID_MICROS
                    )

                    print(f"      ✅ Listing tree created for {shop_name}")

                    # Create shopping product ad in the ad group
                    print(f"      Creating shopping product ad...")
                    ad_resource_name = add_shopping_product_ad(
                        client=client,
                        customer_id=customer_id,
                        ad_group_resource_name=ad_group_resource_name
                    )

                    if not ad_resource_name:
                        print(f"      ⚠️  Warning: Failed to create shopping ad for {shop_name}")

                    shops_processed_successfully.append(shop_name)

                    # Small delay to avoid concurrent modification issues
                    time.sleep(1)

                except Exception as e:
                    error_msg = str(e)
                    print(f"      ❌ Failed to process shop {shop_name}: {error_msg}")
                    shop_errors[shop_name] = error_msg
                    # Continue with next shop instead of failing entire group

            # Mark rows as successful/failed based on their shop
            for row_data in rows_in_group:
                if row_data['shop_name'] in shops_processed_successfully:
                    row_data['row_obj'][COL_LEGACY_STATUS].value = True
                    # Clear error message on success (only if column exists)
                    if len(row_data['row_obj']) > COL_LEGACY_ERROR:
                        row_data['row_obj'][COL_LEGACY_ERROR].value = ""
                else:
                    row_data['row_obj'][COL_LEGACY_STATUS].value = False
                    # Add error message if available (only if column exists)
                    if len(row_data['row_obj']) > COL_LEGACY_ERROR:
                        if row_data['shop_name'] in shop_errors:
                            row_data['row_obj'][COL_LEGACY_ERROR].value = shop_errors[row_data['shop_name']]
                        else:
                            row_data['row_obj'][COL_LEGACY_ERROR].value = "Failed to process shop"

            if len(shops_processed_successfully) > 0:
                successful_groups += 1
                print(f"\n   ✅ GROUP {group_idx} COMPLETED: {len(shops_processed_successfully)}/{len(unique_shops)} shops processed")

            # Save progress periodically
            if file_path and group_idx % 5 == 0:
                print(f"\n   💾 Saving progress...")
                try:
                    workbook.save(file_path)
                except Exception as save_error:
                    print(f"   ⚠️  Failed to save progress: {save_error}")

        except Exception as e:
            error_msg = str(e)
            print(f"\n   ❌ GROUP {group_idx} FAILED: {error_msg}")
            # Mark all rows in this group as failed
            for row_data in rows_in_group:
                row_data['row_obj'][COL_LEGACY_STATUS].value = False
                # Only write error message if column exists
                if len(row_data['row_obj']) > COL_LEGACY_ERROR:
                    row_data['row_obj'][COL_LEGACY_ERROR].value = f"Group failed: {error_msg}"

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Failed to save: {save_error}")

    print(f"\n{'='*70}")
    print(f"INCLUSION SHEET (LEGACY) SUMMARY: {successful_groups}/{total_groups} groups processed successfully")
    print(f"{'='*70}\n")


def _process_single_exclusion_row(
    row_data: dict,
    client: GoogleAdsClient,
    customer_id: str,
    rate_limit_seconds: float
) -> dict:
    """
    Process a single exclusion row (worker function for parallel processing).

    Args:
        row_data: Dict containing row information
        client: Google Ads client
        customer_id: Customer ID
        rate_limit_seconds: Rate limit delay

    Returns:
        Dict with results: {'success': bool, 'error': str or None}
    """
    idx = row_data['idx']
    shop_name = row_data['shop_name']
    cat_uitsluiten = row_data['cat_uitsluiten']
    custom_label_1 = row_data['custom_label_1']

    print(f"\n[Row {idx}] Processing exclusion for shop: {shop_name}")
    print(f"         Category: {cat_uitsluiten}, Custom Label 1: {custom_label_1}")

    # Build campaign name pattern
    campaign_pattern = f"PLA/{cat_uitsluiten}_{custom_label_1}"
    print(f"   Searching for campaign+ad group: {campaign_pattern}")

    # Use combined lookup (saves 1 API call)
    result = get_campaign_and_ad_group_by_pattern(client, customer_id, campaign_pattern)
    if not result:
        print(f"   ❌ Campaign or ad group not found")
        return {
            'success': False,
            'error': f"Campaign not found: {campaign_pattern}"
        }

    campaign = result['campaign']
    ad_group = result['ad_group']

    print(f"   ✅ Found campaign: {campaign['name']} (ID: {campaign['id']})")
    print(f"   ✅ Found ad group: {ad_group['name']} (ID: {ad_group['id']})")

    # Rebuild tree with shop name exclusion
    try:
        rebuild_tree_with_custom_label_3_exclusion(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group['id'],
            shop_name=shop_name,
            default_bid_micros=DEFAULT_BID_MICROS
        )
        print(f"   ✅ SUCCESS - Row {idx} completed")

        # Rate limiting ONLY after successful processing
        if rate_limit_seconds > 0:
            time.sleep(rate_limit_seconds)

        return {'success': True, 'error': None}

    except Exception as e:
        error_msg = str(e)
        print(f"   ❌ Error rebuilding tree: {error_msg}")
        return {
            'success': False,
            'error': f"Error rebuilding tree: {error_msg[:500]}"
        }


def process_uitbreiding_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 5
):
    """
    Process the uitbreiding (extension) sheet - adds shops to existing category campaigns.

    OPTIMIZED VERSION:
    - Groups shops by (maincat, cl1) for batch processing
    - Finds/creates campaign ONCE per group instead of once per shop
    - Reduces API calls significantly

    Excel columns:
    A. Shop name
    B. Shop ID (not used)
    C. maincat (category name)
    D. maincat_id (used as CL4)
    E. custom label 1 (a/b/c)
    F. budget
    G. result (TRUE/FALSE) - updated by script
    H. error message (when status is FALSE)

    Logic:
    1. Group rows by (maincat, custom_label_1)
    2. For each group, find/create campaign ONCE: PLA/{maincat} store_{cl1}
    3. For each shop in group, create ad group: PLA/{shop_name}_{cl1}
    4. Build listing tree for each ad group

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N groups
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING UITBREIDING SHEET: '{SHEET_UITBREIDING}'")
    print(f"(OPTIMIZED: Grouping shops by maincat + cl1)")
    print(f"{'='*70}")

    try:
        sheet = workbook[SHEET_UITBREIDING]
    except KeyError:
        print(f"❌ Sheet '{SHEET_UITBREIDING}' not found in workbook")
        return

    ga_service = client.get_service("GoogleAdsService")

    # =========================================================================
    # STEP 1: Group all rows by (maincat, cl1) for efficient batch processing
    # =========================================================================
    print("\nGrouping rows by (maincat, cl1)...")

    # Structure: {(maincat, cl1): [row_data_dict, ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_UIT_STATUS].value if len(row) > COL_UIT_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_UIT_SHOP_NAME].value
        maincat = row[COL_UIT_MAINCAT].value
        maincat_id = row[COL_UIT_MAINCAT_ID].value
        custom_label_1 = row[COL_UIT_CUSTOM_LABEL_1].value
        budget = row[COL_UIT_BUDGET].value

        # Skip empty rows
        if not shop_name:
            continue

        # Track rows with missing required fields
        if not maincat or not maincat_id or not custom_label_1:
            rows_with_missing_fields.append(idx)
            continue

        group_key = (str(maincat), str(custom_label_1))
        groups[group_key].append({
            'row_idx': idx,
            'shop_name': shop_name,
            'maincat': maincat,
            'maincat_id': maincat_id,
            'custom_label_1': custom_label_1,
            'budget': budget
        })

    # Mark rows with missing fields as errors
    for idx in rows_with_missing_fields:
        print(f"[Row {idx}] ⚠️  Missing required fields, skipping")
        sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = "Missing required fields"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique (maincat, cl1) group(s)")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each group - find/create campaign ONCE per group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0

    for (maincat, cl1), rows_in_group in groups.items():
        groups_processed += 1

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] {maincat} | {cl1}")
        print(f"  Shops in group: {len(rows_in_group)}")
        print(f"{'='*60}")

        # Get metadata from first row
        first_row = rows_in_group[0]
        maincat_id = first_row['maincat_id']
        budget = first_row['budget']

        # Build campaign name - ONCE for entire group
        campaign_name = f"PLA/{maincat} store_{cl1}"
        print(f"\n  Campaign: {campaign_name}")

        try:
            # Step 1: Find or create campaign ONCE for entire group
            escaped_campaign_name = campaign_name.replace("'", "\\'")
            campaign_query = f"""
                SELECT campaign.id, campaign.resource_name, campaign.status
                FROM campaign
                WHERE campaign.name = '{escaped_campaign_name}'
            """

            campaign_results = list(ga_service.search(customer_id=customer_id, query=campaign_query))
            campaign_resource_name = None

            for result in campaign_results:
                if result.campaign.status != client.enums.CampaignStatusEnum.REMOVED:
                    campaign_resource_name = result.campaign.resource_name
                    print(f"  ✅ Found existing campaign")
                    break

            if not campaign_resource_name:
                # Create new campaign
                print(f"  📦 Creating new campaign...")

                # Convert budget from EUR to micros
                try:
                    budget_micros = int(float(budget) * 1_000_000) if budget else 10_000_000
                except (ValueError, TypeError):
                    print(f"     ⚠️  Invalid budget '{budget}', using default 10 EUR")
                    budget_micros = 10_000_000

                # Get bid strategy based on custom label 1
                bid_strategy_resource_name = None
                if cl1 and cl1 in BID_STRATEGY_MAPPING:
                    bid_strategy_name = BID_STRATEGY_MAPPING[cl1]
                    print(f"     Looking up bid strategy: {bid_strategy_name}")
                    bid_strategy_resource_name = get_bid_strategy_by_name(
                        client=client,
                        customer_id=MCC_ACCOUNT_ID,
                        strategy_name=bid_strategy_name
                    )

                # Create campaign
                merchant_center_account_id = 140784594
                budget_name = f"Budget_{campaign_name}"
                first_shop = rows_in_group[0]['shop_name']

                campaign_resource_name = add_standard_shopping_campaign(
                    client=client,
                    customer_id=customer_id,
                    merchant_center_account_id=merchant_center_account_id,
                    campaign_name=campaign_name,
                    budget_name=budget_name,
                    tracking_template="",
                    country="NL",
                    shopid=None,
                    shopname=first_shop,
                    label=cl1,
                    budget=budget_micros,
                    bidding_strategy_resource_name=bid_strategy_resource_name
                )

                if not campaign_resource_name:
                    raise Exception("Failed to create campaign")

                print(f"  ✅ Campaign created")

                # Add negative keyword list to new campaign
                if NEGATIVE_LIST_NAME:
                    print(f"     Adding negative keyword list: {NEGATIVE_LIST_NAME}")
                    enable_negative_list_for_campaign(
                        client=client,
                        customer_id=customer_id,
                        campaign_resource_name=campaign_resource_name,
                        negative_list_name=NEGATIVE_LIST_NAME
                    )

            # Step 2: Process each shop in the group
            print(f"\n  Processing {len(rows_in_group)} shop(s)...")

            for shop_idx, row_data in enumerate(rows_in_group, start=1):
                idx = row_data['row_idx']
                shop_name = row_data['shop_name']
                shop_maincat_id = row_data['maincat_id']

                print(f"\n    [{shop_idx}/{len(rows_in_group)}] {shop_name}")

                try:
                    ad_group_name = f"PLA/{shop_name}_{cl1}"

                    # Look for existing ad group
                    escaped_ad_group_name = ad_group_name.replace("'", "\\'")
                    ad_group_query = f"""
                        SELECT ad_group.id, ad_group.resource_name
                        FROM ad_group
                        WHERE ad_group.campaign = '{campaign_resource_name}'
                        AND ad_group.name = '{escaped_ad_group_name}'
                        AND ad_group.status != 'REMOVED'
                    """

                    ad_group_results = list(ga_service.search(customer_id=customer_id, query=ad_group_query))
                    ad_group_resource_name = None

                    for result in ad_group_results:
                        ad_group_resource_name = result.ad_group.resource_name
                        print(f"      ✅ Found existing ad group")
                        break

                    if not ad_group_resource_name:
                        # Create new ad group
                        print(f"      📦 Creating ad group: {ad_group_name}")
                        ad_group_resource_name, _ = add_shopping_ad_group(
                            client=client,
                            customer_id=customer_id,
                            campaign_resource_name=campaign_resource_name,
                            ad_group_name=ad_group_name,
                            campaign_name=campaign_name
                        )

                        if not ad_group_resource_name:
                            raise Exception("Failed to create ad group")

                        print(f"      ✅ Ad group created")

                    # Build listing tree
                    ad_group_id = ad_group_resource_name.split('/')[-1]

                    build_listing_tree_for_uitbreiding(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        shop_name=shop_name,
                        maincat_id=str(shop_maincat_id),
                        custom_label_1=str(cl1)
                    )

                    # Create shopping product ad
                    add_shopping_product_ad(
                        client=client,
                        customer_id=customer_id,
                        ad_group_resource_name=ad_group_resource_name
                    )

                    # Mark success
                    sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = True
                    sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = ""
                    success_count += 1
                    print(f"      ✅ Row {idx} completed")

                    # Rate limiting between shops
                    time.sleep(1.5)

                except Exception as shop_e:
                    error_msg = str(shop_e)
                    print(f"      ❌ Error: {error_msg[:60]}")

                    # Categorize errors
                    if "CONCURRENT_MODIFICATION" in error_msg:
                        friendly_error = "Concurrent modification (retry needed)"
                    elif "NOT_FOUND" in error_msg.upper():
                        friendly_error = "Resource not found"
                    elif "SUBDIVISION_REQUIRES_OTHERS_CASE" in error_msg:
                        friendly_error = "Tree structure error"
                    else:
                        friendly_error = error_msg[:80]

                    sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = False
                    sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = friendly_error
                    error_count += 1

        except Exception as group_e:
            # Campaign-level error - mark all rows in group as failed
            error_msg = str(group_e)
            print(f"\n  ❌ GROUP FAILED: {error_msg[:80]}")

            for row_data in rows_in_group:
                idx = row_data['row_idx']
                sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = f"Campaign error: {error_msg[:60]}"
                error_count += 1

        # Save periodically (every N groups)
        if file_path and groups_processed % save_interval == 0:
            print(f"\n💾 Saving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"⚠️  Error saving: {save_error}")

        # Delay between groups to avoid concurrent modification
        if groups_processed < total_groups:
            time.sleep(2)

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"UITBREIDING SHEET SUMMARY (OPTIMIZED)")
    print(f"{'='*70}")
    print(f"Total groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Failed: {error_count + len(rows_with_missing_fields)}")
    print(f"{'='*70}\n")


def load_cat_ids_mapping(workbook: openpyxl.Workbook) -> dict:
    """
    Load the cat_ids sheet and create a mapping of maincat_id -> list of deepest_cat values.

    Args:
        workbook: Excel workbook containing cat_ids sheet

    Returns:
        dict: {maincat_id: [deepest_cat1, deepest_cat2, ...]}
    """
    try:
        sheet = workbook[SHEET_CAT_IDS]
    except KeyError:
        print(f"❌ Sheet '{SHEET_CAT_IDS}' not found in workbook")
        return {}

    mapping = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        maincat_id = row[COL_CAT_MAINCAT_ID]
        deepest_cat = row[COL_CAT_DEEPEST_CAT]

        if maincat_id and deepest_cat:
            maincat_id_str = str(maincat_id)
            if maincat_id_str not in mapping:
                mapping[maincat_id_str] = set()
            mapping[maincat_id_str].add(str(deepest_cat))

    # Convert sets to sorted lists
    for key in mapping:
        mapping[key] = sorted(mapping[key])

    print(f"   Loaded {len(mapping)} maincat_id mappings from '{SHEET_CAT_IDS}' sheet")
    return mapping


def add_shop_exclusion_to_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str
):
    """
    Add a shop name as CL3 exclusion to an ad group's listing tree.
    Preserves existing tree structure and adds the shop as a negative CL3 unit.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to exclude (CL3 value)
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Step 1: Read existing tree structure
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    results = list(ga_service.search(customer_id=customer_id, query=query))

    if not results:
        print(f"      ⚠️  No listing tree found in ad group {ad_group_id}")
        return False

    # Find parent for CL3 nodes (by looking at existing CL3 nodes including CL3 OTHERS)
    parent_for_cl3 = None
    existing_cl3_exclusions = set()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        # Check for CL3 nodes (INDEX3) - get parent from any CL3 node
        if lg.case_value.product_custom_attribute.index.name == 'INDEX3':
            value = lg.case_value.product_custom_attribute.value

            if value:
                # This is a specific CL3 value (shop name)
                if criterion.negative:
                    existing_cl3_exclusions.add(value.lower())

            # Get the parent of any CL3 node - this is where we add new CL3 exclusions
            if lg.parent_ad_group_criterion:
                parent_for_cl3 = lg.parent_ad_group_criterion

    if not parent_for_cl3:
        print(f"      ⚠️  No parent for CL3 found in ad group {ad_group_id}")
        return False

    # Check if shop is already excluded
    if shop_name.lower() in existing_cl3_exclusions:
        print(f"      ℹ️  Shop '{shop_name}' already excluded")
        return True

    # Step 2: Add the shop exclusion as a new CL3 negative unit
    dim_cl3_shop = client.get_type("ListingDimensionInfo")
    dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3_shop.product_custom_attribute.value = shop_name

    op = create_listing_group_unit_biddable(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=parent_for_cl3,
        listing_dimension_info=dim_cl3_shop,
        targeting_negative=True,
        cpc_bid_micros=None
    )

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
        print(f"      ✅ Added exclusion: CL3='{shop_name}'")
        return True
    except Exception as e:
        error_msg = str(e)
        if "LISTING_GROUP_ALREADY_EXISTS" in error_msg:
            print(f"      ℹ️  Shop '{shop_name}' already excluded (duplicate)")
            return True
        else:
            print(f"      ❌ Error adding exclusion: {error_msg[:100]}")
            return False


def reverse_exclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    shop_name: str
):
    """
    Remove a shop exclusion from an ad group's listing tree.
    This is the reverse of add_shop_exclusion_to_ad_group - it finds and removes
    the CL3 negative criterion for the specified shop.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to un-exclude (remove CL3 exclusion)

    Returns:
        bool: True if exclusion was removed or didn't exist, False on error
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Step 1: Read existing tree structure to find the CL3 exclusion
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"      ❌ Error reading tree: {e}")
        return False

    if not results:
        print(f"      ⚠️  No listing tree found in ad group {ad_group_name} {ad_group_id}")
        return False

    # Find the CL3 exclusion criterion for this shop
    criterion_to_remove = None
    shop_name_lower = shop_name.lower()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        # Check for CL3 nodes (INDEX3)
        if lg.case_value.product_custom_attribute.index.name == 'INDEX3':
            value = lg.case_value.product_custom_attribute.value

            # Check if this is the shop we want to un-exclude
            if value and value.lower() == shop_name_lower and criterion.negative:
                criterion_to_remove = criterion.resource_name
                break

    if not criterion_to_remove:
        print(f"      ℹ️  Shop '{shop_name}' is not excluded (no exclusion found)")
        return True  # Not an error - shop wasn't excluded

    # Step 2: Remove the exclusion criterion
    op = client.get_type("AdGroupCriterionOperation")
    op.remove = criterion_to_remove

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
        #print(f"      ✅ Removed exclusion: CL3='{shop_name}' from {ad_group_name}")
        return True
    except Exception as e:
        error_msg = str(e)
        print(f"      ❌ Error removing exclusion: {error_msg[:100]}")
        return False


def reverse_exclusion_batch(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    shop_names: list
) -> dict:
    """
    Remove multiple shop exclusions from an ad group's listing tree in one batch.

    This is an optimized version that reads the listing tree ONCE and removes
    all matching shop exclusions in a single mutate operation.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        ad_group_name: Ad group name (for logging)
        shop_names: List of shop names to un-exclude

    Returns:
        dict: {
            'success': list of shop names that were successfully removed,
            'not_found': list of shop names that weren't excluded,
            'errors': list of (shop_name, error_msg) tuples
        }
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    result = {
        'success': [],
        'not_found': [],
        'errors': []
    }

    # Normalize shop names for case-insensitive matching
    shop_names_lower = {name.lower(): name for name in shop_names}

    # Step 1: Read existing tree structure ONCE
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        # All shops failed due to read error
        for shop_name in shop_names:
            result['errors'].append((shop_name, f"Error reading tree: {str(e)[:50]}"))
        return result

    if not results:
        # No listing tree - all shops count as "not found"
        result['not_found'] = list(shop_names)
        return result

    # Step 2: Find ALL CL3 exclusion criteria matching any of our shops
    criteria_to_remove = []  # List of (resource_name, shop_name)
    found_shops = set()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        # Check for CL3 nodes (INDEX3)
        if lg.case_value.product_custom_attribute.index.name == 'INDEX3':
            value = lg.case_value.product_custom_attribute.value

            if value and criterion.negative:
                value_lower = value.lower()
                if value_lower in shop_names_lower:
                    original_shop_name = shop_names_lower[value_lower]
                    criteria_to_remove.append((criterion.resource_name, original_shop_name))
                    found_shops.add(value_lower)

    # Track shops that weren't excluded
    for shop_lower, shop_name in shop_names_lower.items():
        if shop_lower not in found_shops:
            result['not_found'].append(shop_name)

    if not criteria_to_remove:
        # No exclusions to remove
        return result

    # Step 3: Remove all exclusion criteria in ONE batch operation
    operations = []
    for resource_name, shop_name in criteria_to_remove:
        op = client.get_type("AdGroupCriterionOperation")
        op.remove = resource_name
        operations.append((op, shop_name))

    try:
        # Execute batch removal
        agc_service.mutate_ad_group_criteria(
            customer_id=customer_id,
            operations=[op for op, _ in operations]
        )
        # All successful
        for _, shop_name in operations:
            result['success'].append(shop_name)
    except GoogleAdsException as gae:
        # Try to identify which operations failed
        error_msg = str(gae)[:100]
        for _, shop_name in operations:
            result['errors'].append((shop_name, error_msg))
    except Exception as e:
        error_msg = str(e)[:100]
        for _, shop_name in operations:
            result['errors'].append((shop_name, error_msg))

    return result


def prepare_shop_exclusion_operation(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str,
    listing_group_cache: dict = None
) -> tuple:
    """
    Prepare a shop exclusion operation without executing it.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to exclude (CL3 value)
        listing_group_cache: Optional pre-fetched listing group data

    Returns:
        tuple: (operation, status, message)
        - operation: The mutation operation (or None if skip/error)
        - status: 'ready', 'skip', or 'error'
        - message: Description of result
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Use cache if provided, otherwise query
    if listing_group_cache and ad_group_id in listing_group_cache:
        cache_entry = listing_group_cache[ad_group_id]
        parent_for_cl3 = cache_entry.get('parent_for_cl3')
        existing_cl3_exclusions = cache_entry.get('cl3_exclusions', set())
    else:
        # Query listing group structure
        query = f"""
            SELECT
                ad_group_criterion.resource_name,
                ad_group_criterion.listing_group.type,
                ad_group_criterion.listing_group.parent_ad_group_criterion,
                ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
                ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
                ad_group_criterion.negative
            FROM ad_group_criterion
            WHERE ad_group_criterion.ad_group = '{ag_path}'
                AND ad_group_criterion.type = 'LISTING_GROUP'
        """

        try:
            results = list(ga_service.search(customer_id=customer_id, query=query))
        except Exception as e:
            return (None, 'error', f"Query error: {str(e)[:50]}")

        if not results:
            return (None, 'error', "No listing tree found")

        parent_for_cl3 = None
        existing_cl3_exclusions = set()

        for row in results:
            criterion = row.ad_group_criterion
            lg = criterion.listing_group

            # Safely get index name
            try:
                index_name = lg.case_value.product_custom_attribute.index.name
            except (AttributeError, TypeError):
                index_name = None

            # Check for CL3 nodes (INDEX3) - get parent from any CL3 node
            if index_name == 'INDEX3':
                value_str = lg.case_value.product_custom_attribute.value
                if value_str and criterion.negative:
                    existing_cl3_exclusions.add(value_str.lower())
                if lg.parent_ad_group_criterion:
                    parent_for_cl3 = lg.parent_ad_group_criterion

    if not parent_for_cl3:
        return (None, 'error', "No parent for CL3 found")

    # Check if already excluded
    if shop_name.lower() in existing_cl3_exclusions:
        return (None, 'skip', f"Already excluded")

    # Create the operation
    dim_cl3_shop = client.get_type("ListingDimensionInfo")
    dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3_shop.product_custom_attribute.value = shop_name

    op = create_listing_group_unit_biddable(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=parent_for_cl3,
        listing_dimension_info=dim_cl3_shop,
        targeting_negative=True,
        cpc_bid_micros=None
    )

    return (op, 'ready', "Operation prepared")


def execute_exclusion_batch(
    client: GoogleAdsClient,
    customer_id: str,
    operations: list,
    batch_size: int = 10
) -> tuple:
    """
    Execute a batch of exclusion operations.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        operations: List of (operation, ad_group_name, shop_name) tuples
        batch_size: Max operations per API call

    Returns:
        tuple: (success_count, error_count, errors_list)
    """
    if not operations:
        return (0, 0, [])

    agc_service = client.get_service("AdGroupCriterionService")
    success_count = 0
    error_count = 0
    errors = []

    # Process in batches
    for i in range(0, len(operations), batch_size):
        batch = operations[i:i + batch_size]
        ops = [item[0] for item in batch]

        try:
            agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
            success_count += len(batch)
            for _, ag_name, shop_name in batch:
                print(f"      ✅ {ag_name}: excluded '{shop_name}'")
        except Exception as e:
            error_msg = str(e)
            if "LISTING_GROUP_ALREADY_EXISTS" in error_msg:
                # Some might already exist, count as success
                success_count += len(batch)
                print(f"      ℹ️  Batch contained duplicates (already excluded)")
            else:
                # Batch failed - try individual operations
                print(f"      ⚠️  Batch failed, retrying individually...")
                for op, ag_name, shop_name in batch:
                    try:
                        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
                        success_count += 1
                        print(f"      ✅ {ag_name}: excluded '{shop_name}'")
                    except Exception as ind_e:
                        ind_error = str(ind_e)
                        if "LISTING_GROUP_ALREADY_EXISTS" in ind_error:
                            success_count += 1
                            print(f"      ℹ️  {ag_name}: already excluded")
                        else:
                            error_count += 1
                            errors.append(f"{ag_name}: {ind_error[:50]}")
                            print(f"      ❌ {ag_name}: {ind_error[:50]}")
                    time.sleep(0.1)

        # Small delay between batches
        if i + batch_size < len(operations):
            time.sleep(0.2)

    return (success_count, error_count, errors)


def add_shop_exclusions_batch(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    shop_names: list
) -> dict:
    """
    Add multiple shop exclusions to an ad group's listing tree in one batch.

    This is an optimized version that reads the listing tree ONCE and adds
    all exclusions in a single mutate operation.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        ad_group_name: Ad group name (for logging)
        shop_names: List of shop names to exclude

    Returns:
        dict: {
            'success': list of shop names that were successfully excluded,
            'already_excluded': list of shop names that were already excluded,
            'errors': list of (shop_name, error_msg) tuples
        }
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    result = {
        'success': [],
        'already_excluded': [],
        'errors': []
    }

    # Normalize shop names for case-insensitive matching
    shop_names_lower = {name.lower(): name for name in shop_names}

    # Step 1: Read existing tree structure ONCE
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        # All shops failed due to read error
        for shop_name in shop_names:
            result['errors'].append((shop_name, f"Error reading tree: {str(e)[:50]}"))
        return result

    if not results:
        for shop_name in shop_names:
            result['errors'].append((shop_name, "No listing tree found"))
        return result

    # Step 2: Find parent for CL3 and existing exclusions
    parent_for_cl3 = None
    existing_cl3_exclusions = set()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        try:
            index_name = lg.case_value.product_custom_attribute.index.name
        except (AttributeError, TypeError):
            index_name = None

        if index_name == 'INDEX3':
            value_str = lg.case_value.product_custom_attribute.value
            if value_str and criterion.negative:
                existing_cl3_exclusions.add(value_str.lower())
            if lg.parent_ad_group_criterion:
                parent_for_cl3 = lg.parent_ad_group_criterion

    if not parent_for_cl3:
        for shop_name in shop_names:
            result['errors'].append((shop_name, "No parent for CL3 found"))
        return result

    # Step 3: Determine which shops to add vs skip
    operations = []
    for shop_lower, shop_name in shop_names_lower.items():
        if shop_lower in existing_cl3_exclusions:
            result['already_excluded'].append(shop_name)
        else:
            # Create operation for this shop
            dim_cl3_shop = client.get_type("ListingDimensionInfo")
            dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            dim_cl3_shop.product_custom_attribute.value = shop_name

            op = create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=parent_for_cl3,
                listing_dimension_info=dim_cl3_shop,
                targeting_negative=True,
                cpc_bid_micros=None
            )
            operations.append((op, shop_name))

    if not operations:
        # All shops were already excluded
        return result

    # Step 4: Execute batch
    try:
        ops = [op for op, _ in operations]
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
        for _, shop_name in operations:
            result['success'].append(shop_name)
    except GoogleAdsException as gae:
        error_str = str(gae)
        if "LISTING_GROUP_ALREADY_EXISTS" in error_str:
            # Some were duplicates - try individually
            for op, shop_name in operations:
                try:
                    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
                    result['success'].append(shop_name)
                except Exception as ind_e:
                    ind_error = str(ind_e)
                    if "LISTING_GROUP_ALREADY_EXISTS" in ind_error:
                        result['already_excluded'].append(shop_name)
                    else:
                        result['errors'].append((shop_name, ind_error[:50]))
        else:
            # Batch failed - try individually
            for op, shop_name in operations:
                try:
                    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
                    result['success'].append(shop_name)
                except Exception as ind_e:
                    ind_error = str(ind_e)
                    if "LISTING_GROUP_ALREADY_EXISTS" in ind_error:
                        result['already_excluded'].append(shop_name)
                    else:
                        result['errors'].append((shop_name, ind_error[:50]))
    except Exception as e:
        error_msg = str(e)[:100]
        for _, shop_name in operations:
            result['errors'].append((shop_name, error_msg))

    return result


def prefetch_pla_campaigns_and_ad_groups(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_prefix: str = "PLA/"
) -> dict:
    """
    Pre-fetch all PLA campaigns and their ad groups in a single query.

    Returns a dict structured as:
    {
        'campaign_name': {
            'resource_name': 'customers/xxx/campaigns/yyy',
            'ad_groups': [
                {'id': 123, 'name': 'ag_name', 'resource_name': 'customers/xxx/adGroups/zzz'},
                ...
            ]
        },
        ...
    }
    """
    print(f"\n📥 Pre-fetching campaigns and ad groups (prefix: {campaign_prefix})...")

    ga_service = client.get_service("GoogleAdsService")

    escaped_prefix = campaign_prefix.replace("'", "\\'")
    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name
        FROM ad_group
        WHERE campaign.name LIKE '{escaped_prefix}%'
        AND campaign.status != 'REMOVED'
        AND ad_group.status != 'REMOVED'
        ORDER BY campaign.name, ad_group.name
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"❌ Error pre-fetching: {e}")
        return {}

    # Build the cache structure
    cache = {}
    for row in results:
        campaign_name = row.campaign.name
        if campaign_name not in cache:
            cache[campaign_name] = {
                'resource_name': row.campaign.resource_name,
                'ad_groups': []
            }
        cache[campaign_name]['ad_groups'].append({
            'id': row.ad_group.id,
            'name': row.ad_group.name,
            'resource_name': row.ad_group.resource_name
        })

    total_campaigns = len(cache)
    total_ad_groups = sum(len(c['ad_groups']) for c in cache.values())
    print(f"✅ Cached {total_campaigns} campaigns with {total_ad_groups} ad groups\n")

    return cache


def process_exclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10
):
    """
    Process the 'uitsluiten' (exclusion) sheet - V2 with cat_ids mapping.

    OPTIMIZED VERSION (V2):
    - Groups shops by (maincat_id, cl1) for batch processing
    - Reads each ad group's listing tree ONCE per group instead of once per shop
    - Pre-fetches all PLA campaigns and ad groups upfront
    - Uses batch mutations for faster processing

    Excel columns (uitsluiten):
    A. Shop name - shop to exclude
    B. Shop ID (not used)
    C. maincat - category name
    D. maincat_id - used to look up deepest_cats
    E. custom label 1 (a/b/c)
    F. result (TRUE/FALSE) - updated by script
    G. error message (when status is FALSE)

    Logic:
    1. Group all rows by (maincat_id, custom_label_1)
    2. For each group, look up deepest_cats ONCE
    3. For each deepest_cat, find campaign PLA/{deepest_cat}_{cl1}
    4. For each ad group, add ALL shop exclusions in one batch

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N groups
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING EXCLUSION SHEET V2: '{SHEET_EXCLUSION}'")
    print(f"(OPTIMIZED: Grouping shops by maincat_id + cl1)")
    print(f"{'='*70}")

    # Load cat_ids mapping
    print("\nLoading cat_ids mapping...")
    cat_ids_mapping = load_cat_ids_mapping(workbook)
    if not cat_ids_mapping:
        print("❌ No cat_ids mapping loaded, cannot process exclusions")
        return

    try:
        sheet = workbook[SHEET_EXCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_EXCLUSION}' not found in workbook")
        return

    # Pre-fetch all PLA campaigns and ad groups
    print("\nPre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    # =========================================================================
    # STEP 1: Group all rows by (maincat_id, cl1) for efficient batch processing
    # =========================================================================
    print("\nGrouping rows by (maincat_id, cl1)...")

    # Structure: {(maincat_id, cl1): [(row_idx, shop_name), ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_EX_STATUS].value if len(row) > COL_EX_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_EX_SHOP_NAME].value
        maincat_id = row[COL_EX_MAINCAT_ID].value
        custom_label_1 = row[COL_EX_CUSTOM_LABEL_1].value

        # Skip empty rows
        if not shop_name:
            continue

        # Track rows with missing required fields
        if not maincat_id or not custom_label_1:
            rows_with_missing_fields.append(idx)
            continue

        maincat_id_str = str(maincat_id)
        cl1_str = str(custom_label_1)
        groups[(maincat_id_str, cl1_str)].append((idx, shop_name))

    # Mark rows with missing fields as errors
    for idx in rows_with_missing_fields:
        print(f"[Row {idx}] ⚠️  Missing required fields, skipping")
        sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = "Missing required fields"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique (maincat_id, cl1) group(s)")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each group - fetch campaigns/ad groups ONCE per group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0

    for (maincat_id_str, cl1_str), rows in groups.items():
        groups_processed += 1
        shop_names = [shop_name for _, shop_name in rows]
        row_indices = [idx for idx, _ in rows]

        # For CL3 targeting, split shop_name at | and use first part
        # e.g. "Hbm-machines.com|NL" becomes "Hbm-machines.com"
        shop_names_for_targeting = [name.split('|')[0] if '|' in name else name for name in shop_names]
        # Create mapping from targeting name back to original name(s)
        targeting_to_original = {}
        for orig, tgt in zip(shop_names, shop_names_for_targeting):
            if tgt not in targeting_to_original:
                targeting_to_original[tgt] = []
            targeting_to_original[tgt].append(orig)

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] maincat_id={maincat_id_str}, cl1={cl1_str}")
        print(f"  Shops to process: {len(shop_names)}")
        print(f"  Shop names: {', '.join(shop_names[:5])}{'...' if len(shop_names) > 5 else ''}")
        # Show if any names were split
        split_names = [(orig, tgt) for orig, tgt in zip(shop_names, shop_names_for_targeting) if orig != tgt]
        if split_names:
            print(f"  CL3 targeting (split): {', '.join([f'{tgt} (from {orig})' for orig, tgt in split_names[:3]])}{'...' if len(split_names) > 3 else ''}")

        # Look up deepest_cats for this maincat_id ONCE for the entire group
        deepest_cats = cat_ids_mapping.get(maincat_id_str, [])
        if not deepest_cats:
            print(f"  ⚠️  No deepest_cats found for maincat_id={maincat_id_str}")
            # Mark all rows in this group as failed
            for idx in row_indices:
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = f"No deepest_cats for maincat_id={maincat_id_str}"
                error_count += 1
            continue

        print(f"  Found {len(deepest_cats)} deepest_cat(s)")

        # Track results per shop
        shop_results = {shop: {'success': 0, 'already_excluded': 0, 'errors': []} for shop in shop_names}
        campaigns_found = 0
        total_exclusions_added = 0

        # Process each deepest_cat ONCE for all shops in this group
        for deepest_cat in deepest_cats:
            campaign_name = f"PLA/{deepest_cat}_{cl1_str}"
            campaign_data = campaign_cache.get(campaign_name)

            if campaign_data:
                campaigns_found += 1
                ad_groups = campaign_data['ad_groups']

                for ag in ad_groups:
                    ag_id = str(ag['id'])
                    ag_name = ag['name']

                    # Retry logic for connection errors
                    max_retries = 3
                    retry_delay = 2

                    for attempt in range(max_retries):
                        try:
                            # Call batch function with targeting names (split at |)
                            # Use unique targeting names to avoid duplicates
                            unique_targeting_names = list(set(shop_names_for_targeting))
                            result = add_shop_exclusions_batch(
                                client=client,
                                customer_id=customer_id,
                                ad_group_id=ag_id,
                                ad_group_name=ag_name,
                                shop_names=unique_targeting_names
                            )

                            # Aggregate results - map targeting names back to original names
                            for targeting_name in result['success']:
                                # Update all original names that map to this targeting name
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['success'] += 1
                                        total_exclusions_added += 1
                            for targeting_name in result['already_excluded']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['already_excluded'] += 1
                            for targeting_name, error in result['errors']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['errors'].append(f"{ag_name}: {error}")

                            break  # Success, exit retry loop

                        except Exception as e:
                            error_str = str(e)
                            if "failed to connect" in error_str.lower() or "unavailable" in error_str.lower():
                                if attempt < max_retries - 1:
                                    print(f"    ⚠️  Connection error, retrying in {retry_delay}s...")
                                    time.sleep(retry_delay)
                                    retry_delay *= 2
                                    continue
                            # Non-retryable error or max retries reached
                            error_msg = str(e)[:50]
                            print(f"    ❌ {ag_name}: {error_msg}")
                            for shop in shop_names:
                                shop_results[shop]['errors'].append(f"{ag_name}: {error_msg}")
                            break

                    # Rate limiting delay after each ad group (not each shop!)
                    time.sleep(0.3)

        print(f"  Found {campaigns_found} campaign(s), added {total_exclusions_added} exclusion(s) total")

        # =========================================================================
        # STEP 3: Update row statuses based on results
        # =========================================================================
        for idx, shop_name in rows:
            result = shop_results[shop_name]

            # Consider success if: at least one exclusion added OR already excluded
            # AND no errors occurred
            has_errors = len(result['errors']) > 0
            has_activity = result['success'] > 0 or result['already_excluded'] > 0

            if campaigns_found == 0:
                # No campaigns found at all - this is an error
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = f"No campaigns found for maincat_id={maincat_id_str}"
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ No campaigns")
            elif has_errors:
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
                error_summary = "; ".join(result['errors'][:3])
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = error_summary[:100]
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ {len(result['errors'])} error(s)")
            else:
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = True
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = ""
                success_count += 1
                print(f"    Row {idx} ({shop_name}): ✅ added={result['success']}, already={result['already_excluded']}")

        # Save periodically (every N groups)
        if file_path and groups_processed % save_interval == 0:
            print(f"\n💾 Saving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"EXCLUSION SHEET V2 SUMMARY (OPTIMIZED)")
    print(f"{'='*70}")
    print(f"Total groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Failed: {error_count + len(rows_with_missing_fields)}")
    print(f"{'='*70}\n")


def process_exclusion_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str,
    save_interval: int = 10
):
    """
    Process the 'uitsluiten' (exclusion) sheet with GROUPED PROCESSING.

    Groups rows by campaign (cat_uitsluiten + custom_label_1) and collects all
    shops to exclude for each campaign. Then rebuilds each campaign's tree once
    with all shop exclusions.

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file for saving
        save_interval: Save workbook every N campaign groups (default: 10)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING EXCLUSION SHEET: '{SHEET_EXCLUSION}' (GROUPED MODE)")
    print(f"{'='*70}")
    print(f"  Strategy: Group rows by campaign, apply all shop exclusions at once")
    print(f"  Save interval: Every {save_interval} campaign groups")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[SHEET_EXCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_EXCLUSION}' not found in workbook")
        return

    # Step 1: Group rows by campaign and collect shops
    print("Step 1: Grouping rows by campaign...")
    campaign_groups = defaultdict(lambda: {'rows': [], 'shops': set()})

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if row has enough columns
        if len(row) <= COL_EX_CUSTOM_LABEL_1:
            print(f"⚠️  Row {idx}: Not enough columns (has {len(row)}, needs at least {COL_EX_CUSTOM_LABEL_1 + 1}). Skipping.")
            continue

        # Skip rows that already have a status
        status_cell = sheet.cell(row=idx, column=COL_EX_STATUS + 1)  # +1 because openpyxl is 1-indexed
        if status_cell.value is not None and status_cell.value != '':
            continue

        # Extract values safely
        try:
            shop_name = row[COL_EX_SHOP_NAME].value
            cat_uitsluiten = row[COL_EX_CAT_UITSLUITEN].value
            diepste_cat_id = row[COL_EX_DIEPSTE_CAT_ID].value
            custom_label_1 = row[COL_EX_CUSTOM_LABEL_1].value
        except IndexError as e:
            print(f"⚠️  Row {idx}: Column access error: {e}. Skipping.")
            continue

        # Validate required fields
        if not shop_name or not cat_uitsluiten or not custom_label_1 or not diepste_cat_id:
            sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = "Missing required fields"
            continue

        # Group key: (cat_uitsluiten, custom_label_1)
        group_key = (cat_uitsluiten, str(custom_label_1))

        # Add row and shop to group - store row number, not row tuple
        campaign_groups[group_key]['rows'].append({
            'idx': idx,
            'row_number': idx
        })
        campaign_groups[group_key]['shops'].add(str(shop_name))
        # Store diepste_cat_id (should be same for all rows in group)
        campaign_groups[group_key]['diepste_cat_id'] = str(diepste_cat_id)

    print(f"Found {len(campaign_groups)} campaign group(s) to process")
    print(f"Total rows: {sum(len(g['rows']) for g in campaign_groups.values())}\n")

    if len(campaign_groups) == 0:
        print("✅ No campaign groups to process")
        return

    # Step 2: Process each campaign group
    print("="*70)
    print("Step 2: Processing campaign groups...")
    print("="*70)

    success_count = 0
    fail_count = 0
    groups_processed = 0

    for i, (group_key, group_data) in enumerate(campaign_groups.items(), 1):
        try:
            cat_uitsluiten, custom_label_1 = group_key
        except (ValueError, TypeError) as e:
            print(f"\n❌ ERROR unpacking group_key: {group_key}")
            print(f"   Error: {e}")
            print(f"   Skipping this group...")
            continue

        rows = group_data['rows']
        shops = sorted(group_data['shops'])
        diepste_cat_id = group_data.get('diepste_cat_id')

        campaign_pattern = f"PLA/{cat_uitsluiten}_{custom_label_1}"

        print(f"\n{'─'*70}")
        print(f"GROUP {i}/{len(campaign_groups)}: {campaign_pattern}")
        print(f"{'─'*70}")
        print(f"   Rows in group: {len(rows)}")
        print(f"   Diepste cat ID (CL0): {diepste_cat_id}")
        print(f"   Shops to exclude: {len(shops)}")
        print(f"   Shop names: {', '.join(shops)}")

        try:
            # Find campaign and ad group
            result = get_campaign_and_ad_group_by_pattern(client, customer_id, campaign_pattern)

            if not result:
                print(f"   ❌ Campaign not found")
                # Mark all rows in group as NOT_FOUND
                for row_info in rows:
                    row_num = row_info['row_number']
                    sheet.cell(row=row_num, column=COL_EX_STATUS + 1).value = False
                    sheet.cell(row=row_num, column=COL_EX_ERROR + 1).value = "Campaign not found"
                    fail_count += 1
                continue

            print(f"   ✅ Found: Campaign ID {result['campaign']['id']}, Ad Group ID {result['ad_group']['id']}")

            # Rebuild tree with all shop exclusions and required CL0 targeting
            rebuild_tree_with_shop_exclusions(
                client,
                customer_id,
                result['ad_group']['id'],
                shop_names=shops,  # Pass all shops for this campaign
                required_cl0_value=diepste_cat_id  # Required CL0 from Excel
            )

            # Mark all rows in group as SUCCESS
            for row_info in rows:
                row_num = row_info['row_number']
                sheet.cell(row=row_num, column=COL_EX_STATUS + 1).value = True
                sheet.cell(row=row_num, column=COL_EX_ERROR + 1).value = ""  # Clear error message
                success_count += 1

            groups_processed += 1
            print(f"   ✅ SUCCESS - Tree rebuilt with {len(shops)} shop exclusion(s)")

        except Exception as e:
            print(f"   ❌ ERROR: {e}")
            # Mark all rows in group as ERROR
            # Create brief, user-friendly error message
            error_str = str(e)

            # Shorten common error types
            if "SUBDIVISION_REQUIRES_OTHERS_CASE" in error_str:
                error_msg = "Tree structure error: missing OTHERS case"
            elif "LISTING_GROUP_SUBDIVISION_REQUIRES_OTHERS_CASE" in error_str:
                error_msg = "Tree structure error: missing OTHERS case"
            elif "CONCURRENT_MODIFICATION" in error_str:
                error_msg = "Concurrent modification (retry needed)"
            elif "NOT_FOUND" in error_str or "not found" in error_str.lower():
                error_msg = "Resource not found"
            elif "INVALID_ARGUMENT" in error_str:
                error_msg = "Invalid argument in API call"
            elif "PERMISSION_DENIED" in error_str:
                error_msg = "Permission denied"
            elif "Could not find CL0" in error_str or "Could not find CL1" in error_str:
                error_msg = error_str[:80]  # Keep this one as-is, it's informative
            else:
                # Generic error - truncate but keep key info
                error_msg = error_str[:80] if len(error_str) > 80 else error_str

            for row_info in rows:
                row_num = row_info['row_number']
                sheet.cell(row=row_num, column=COL_EX_STATUS + 1).value = False
                sheet.cell(row=row_num, column=COL_EX_ERROR + 1).value = error_msg
                fail_count += 1

        # Save every N groups
        if i % save_interval == 0:
            print(f"\n   💾 Saving progress... ({i}/{len(campaign_groups)} groups processed)")
            try:
                workbook.save(file_path)
                print(f"   ✅ Progress saved successfully")
            except Exception as save_error:
                print(f"   ⚠️  Error saving file: {save_error}")

    # Final save
    print(f"\n   💾 Final save...")
    try:
        workbook.save(file_path)
        print(f"   ✅ Final save successful")
    except Exception as save_error:
        print(f"   ⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"EXCLUSION SHEET SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaign groups processed: {len(campaign_groups)}")
    print(f"✅ Groups successful: {groups_processed}")
    print(f"❌ Groups failed: {len(campaign_groups) - groups_processed}")
    print(f"✅ Total rows marked success: {success_count}")
    print(f"❌ Total rows marked failed: {fail_count}")
    print(f"{'='*70}\n")


# ============================================================================
# CL1 VALIDATION FUNCTIONS (Legacy-based)
# ============================================================================

def validate_cl1_targeting_for_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    dry_run: bool = False
) -> dict:
    """
    Validate and fix CL1 (Custom Label 1) targeting for an ad group.

    Checks if the product listing tree has a product group targeting the custom value
    (a, b, or c) that matches the ad group name suffix (_a, _b, or _c).

    If the correct CL1 targeting is missing, it will be added while preserving
    existing exclusions and tree structure.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        ad_group_name: Ad group name (to extract suffix)
        dry_run: If True, only report issues without making changes

    Returns:
        dict with keys:
            - 'status': 'ok', 'fixed', 'skipped', 'error'
            - 'message': Description of what was done/found
            - 'required_cl1': The required CL1 value from ad group name
            - 'existing_cl1': The existing CL1 value(s) found in tree
    """
    result = {
        'status': 'ok',
        'message': '',
        'required_cl1': None,
        'existing_cl1': []
    }

    # Step 1: Extract required CL1 from ad group name suffix
    required_cl1 = None
    for suffix in ['_a', '_b', '_c']:
        if ad_group_name.endswith(suffix):
            required_cl1 = suffix[1:]  # Remove underscore: "_a" → "a"
            break

    if not required_cl1:
        result['status'] = 'skipped'
        result['message'] = f"Ad group name '{ad_group_name}' does not end with _a, _b, or _c"
        return result

    result['required_cl1'] = required_cl1

    # Step 2: Query existing listing tree
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        rows = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        result['status'] = 'error'
        result['message'] = f"Error querying listing tree: {str(e)[:100]}"
        return result

    if not rows:
        result['status'] = 'skipped'
        result['message'] = "No listing tree found in ad group"
        return result

    # Step 3: Analyze the tree structure
    # Find CL4 subdivision (parent for CL1 nodes) and existing CL1 targeting
    cl4_subdivision_resource = None
    existing_cl1_values = []  # Positive CL1 targets
    cl1_others_exists = False
    existing_bid = DEFAULT_BID_MICROS

    for row in rows:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group
        case_value = lg.case_value

        # Check for custom attribute nodes
        if case_value.product_custom_attribute:
            index = case_value.product_custom_attribute.index.name
            value = case_value.product_custom_attribute.value

            # Track CL4 subdivisions (these are parents for CL1 nodes)
            if index == 'INDEX4' and lg.type.name == 'SUBDIVISION':
                cl4_subdivision_resource = criterion.resource_name

            # Track existing CL1 targeting
            if index == 'INDEX1':
                if value:
                    # Specific CL1 value
                    if not criterion.negative:
                        existing_cl1_values.append(value)
                        # Capture bid from positive CL1 units
                        if criterion.cpc_bid_micros:
                            existing_bid = criterion.cpc_bid_micros
                else:
                    # CL1 OTHERS case
                    cl1_others_exists = True

    result['existing_cl1'] = existing_cl1_values

    # Step 4: Check if required CL1 is already targeted
    if required_cl1 in existing_cl1_values:
        result['status'] = 'ok'
        result['message'] = f"CL1='{required_cl1}' already targeted correctly"
        return result

    # CL1 is missing - need to add it
    if not cl4_subdivision_resource:
        result['status'] = 'error'
        result['message'] = "Could not find CL4 subdivision (maincat) to add CL1 under"
        return result

    if dry_run:
        result['status'] = 'fixed'
        result['message'] = f"[DRY RUN] Would add CL1='{required_cl1}' targeting (existing: {existing_cl1_values or 'none'})"
        return result

    # Step 5: Add the missing CL1 targeting
    ops = []

    # If CL1 OTHERS doesn't exist, we need to add it first
    # (Every subdivision must have an OTHERS case)
    if not cl1_others_exists:
        dim_cl1_others = client.get_type("ListingDimensionInfo")
        dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        # Don't set value - this is the OTHERS case

        ops.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl4_subdivision_resource,
                listing_dimension_info=dim_cl1_others,
                targeting_negative=True,  # OTHERS is always negative
                cpc_bid_micros=None
            )
        )

    # Add the positive CL1 target
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    dim_cl1.product_custom_attribute.value = required_cl1

    ops.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl4_subdivision_resource,
            listing_dimension_info=dim_cl1,
            targeting_negative=False,  # Positive target
            cpc_bid_micros=existing_bid  # Use existing bid or default
        )
    )

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
        result['status'] = 'fixed'
        result['message'] = f"Added CL1='{required_cl1}' targeting (bid: {existing_bid/1_000_000:.2f}€)"
        if not cl1_others_exists:
            result['message'] += " + added CL1 OTHERS"
        return result
    except Exception as e:
        error_msg = str(e)
        if "LISTING_GROUP_ALREADY_EXISTS" in error_msg:
            result['status'] = 'ok'
            result['message'] = f"CL1='{required_cl1}' already exists (concurrent update)"
            return result
        result['status'] = 'error'
        result['message'] = f"Error adding CL1 targeting: {error_msg[:100]}"
        return result


def validate_cl1_targeting_for_campaigns(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_name_pattern: str = None,
    dry_run: bool = False
) -> dict:
    """
    Validate and fix CL1 targeting for all ad groups in matching campaigns.

    Iterates through all campaigns (optionally filtered by name pattern) and
    validates that each ad group's product listing tree targets the correct
    CL1 value based on the ad group name suffix (_a, _b, or _c).

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_name_pattern: Optional pattern to filter campaigns (uses LIKE)
                              e.g., "PLA/%" for all PLA campaigns
        dry_run: If True, only report issues without making changes

    Returns:
        dict with summary statistics and details
    """
    print(f"\n{'='*70}")
    print("CL1 TARGETING VALIDATION")
    print(f"{'='*70}")
    print(f"Customer ID: {customer_id}")
    print(f"Campaign filter: {campaign_name_pattern or '(all campaigns)'}")
    print(f"Dry run: {dry_run}")
    print(f"{'='*70}\n")

    ga_service = client.get_service("GoogleAdsService")

    # Step 1: Query campaigns and ad groups
    where_clause = "campaign.status != 'REMOVED' AND ad_group.status != 'REMOVED'"
    if campaign_name_pattern:
        # Escape single quotes in pattern
        escaped_pattern = campaign_name_pattern.replace("'", "\\'")
        where_clause += f" AND campaign.name LIKE '{escaped_pattern}'"

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name
        FROM ad_group
        WHERE {where_clause}
        ORDER BY campaign.name, ad_group.name
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"❌ Error querying campaigns/ad groups: {e}")
        return {'error': str(e)}

    print(f"Found {len(results)} ad group(s) to validate\n")

    # Step 2: Process each ad group
    stats = {
        'total': len(results),
        'ok': 0,
        'fixed': 0,
        'skipped': 0,
        'error': 0,
        'details': []
    }

    current_campaign = None

    for row in results:
        campaign_name = row.campaign.name
        ag_id = row.ad_group.id
        ag_name = row.ad_group.name

        # Print campaign header when it changes
        if campaign_name != current_campaign:
            current_campaign = campaign_name
            print(f"\n📁 Campaign: {campaign_name}")

        # Validate this ad group
        result = validate_cl1_targeting_for_ad_group(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ag_id),
            ad_group_name=ag_name,
            dry_run=dry_run
        )

        # Update stats
        stats[result['status']] += 1

        # Log result
        status_icon = {
            'ok': '✅',
            'fixed': '🔧',
            'skipped': '⏭️',
            'error': '❌'
        }.get(result['status'], '❓')

        print(f"   {status_icon} {ag_name}: {result['message']}")

        # Store details for reporting
        stats['details'].append({
            'campaign': campaign_name,
            'ad_group': ag_name,
            'ad_group_id': ag_id,
            **result
        })

        # Rate limiting
        if result['status'] == 'fixed' and not dry_run:
            time.sleep(0.5)

    # Print summary
    print(f"\n{'='*70}")
    print("CL1 VALIDATION SUMMARY")
    print(f"{'='*70}")
    print(f"Total ad groups: {stats['total']}")
    print(f"✅ Already correct: {stats['ok']}")
    print(f"🔧 Fixed: {stats['fixed']}")
    print(f"⏭️  Skipped (no _a/_b/_c suffix): {stats['skipped']}")
    print(f"❌ Errors: {stats['error']}")
    print(f"{'='*70}\n")

    # Write ad groups that need fixing to xlsx file
    to_fix = [d for d in stats['details'] if d['status'] == 'fixed']
    if to_fix:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ad Groups to Fix"
        ws.append(["Campaign Name", "Ad Group Name"])
        for item in to_fix:
            ws.append([item['campaign'], item['ad_group']])
        output_file = r"C:\Users\JoepvanSchagen\Downloads\Python\scripts_def\DMA+\cl1_validation_to_fix.xlsx"
        wb.save(output_file)
        print(f"📄 Wrote {len(to_fix)} ad group(s) to {output_file}")

    return stats


def process_reverse_exclusion_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10,
    sheet_name: str = "verwijderen"
):
    """
    Process a sheet to REMOVE shop exclusions (reverse of exclusion).

    OPTIMIZED VERSION: Groups shops by (maincat_id, cl1) and processes them together,
    reading each ad group's listing tree only ONCE per group instead of once per shop.

    This function reads a sheet with shop/category data and removes the CL3 exclusion
    for each shop, effectively un-excluding them from the campaigns.

    Uses the same cat_ids mapping approach as process_exclusion_sheet_v2:
    1. Get maincat_id from input row
    2. Look up all deepest_cats for that maincat_id in cat_ids sheet
    3. For each deepest_cat, find campaign PLA/{deepest_cat}_{cl1}
    4. Remove shop exclusions from all ad groups in that campaign (batched)

    Excel columns (same structure as uitsluiten):
    A. Shop name - shop to un-exclude
    B. Shop ID - (not used, for reference)
    C. maincat - Category name (for reference)
    D. maincat_id - Maincat ID (used to look up deepest_cats via cat_ids sheet)
    E. custom label 1 - Custom label 1 value (for campaign matching)
    F. Status - Will be updated with TRUE/FALSE
    G. Error - Will be updated with error messages

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to save progress
        save_interval: Save every N groups processed
        sheet_name: Name of the sheet to process (default: "verwijderen")
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING REVERSE EXCLUSION SHEET: '{sheet_name}'")
    print(f"(OPTIMIZED: Grouping shops by maincat_id + cl1)")
    print(f"{'='*70}")

    # Load cat_ids mapping (same as process_exclusion_sheet_v2)
    print("\nLoading cat_ids mapping...")
    cat_ids_mapping = load_cat_ids_mapping(workbook)
    if not cat_ids_mapping:
        print("❌ No cat_ids mapping loaded, cannot process reverse exclusions")
        return

    if sheet_name not in workbook.sheetnames:
        print(f"⚠️  Sheet '{sheet_name}' not found in workbook")
        print(f"   Available sheets: {workbook.sheetnames}")
        return

    sheet = workbook[sheet_name]

    # Column indices (0-based) - same structure as uitsluiten sheet
    COL_SHOP_NAME = 0      # A: Shop name
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat (category name, for reference)
    COL_MAINCAT_ID = 3     # D: maincat_id (used to look up deepest_cats)
    COL_CL1 = 4            # E: custom label 1
    COL_STATUS = 5         # F: Status
    COL_ERROR = 6          # G: Error

    # Pre-fetch campaigns cache
    print("\nPre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    # =========================================================================
    # STEP 1: Group all rows by (maincat_id, cl1) for efficient batch processing
    # =========================================================================
    print("\nGrouping rows by (maincat_id, cl1)...")

    # Structure: {(maincat_id, cl1): [(row_idx, shop_name), ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []  # Track rows with missing required fields

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_STATUS].value if len(row) > COL_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value
        maincat_id = row[COL_MAINCAT_ID].value
        custom_label_1 = row[COL_CL1].value

        # Skip empty rows
        if not shop_name:
            continue

        # Track rows with missing required fields
        if not maincat_id or not custom_label_1:
            rows_with_missing_fields.append(idx)
            continue

        maincat_id_str = str(maincat_id)
        cl1_str = str(custom_label_1)
        groups[(maincat_id_str, cl1_str)].append((idx, shop_name))

    # Mark rows with missing fields as errors
    for idx in rows_with_missing_fields:
        print(f"[Row {idx}] ⚠️  Missing required fields, skipping")
        sheet.cell(row=idx, column=COL_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_ERROR + 1).value = "Missing required fields"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique (maincat_id, cl1) group(s)")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each group - fetch campaigns/ad groups ONCE per group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0

    for (maincat_id_str, cl1_str), rows in groups.items():
        groups_processed += 1
        shop_names = [shop_name for _, shop_name in rows]
        row_indices = [idx for idx, _ in rows]

        # For CL3 targeting, split shop_name at | and use first part
        # e.g. "Hbm-machines.com|NL" becomes "Hbm-machines.com"
        shop_names_for_targeting = [name.split('|')[0] if '|' in name else name for name in shop_names]
        # Create mapping from targeting name back to original name(s)
        targeting_to_original = {}
        for orig, tgt in zip(shop_names, shop_names_for_targeting):
            if tgt not in targeting_to_original:
                targeting_to_original[tgt] = []
            targeting_to_original[tgt].append(orig)

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] maincat_id={maincat_id_str}, cl1={cl1_str}")
        print(f"  Shops to process: {len(shop_names)}")
        print(f"  Shop names: {', '.join(shop_names[:5])}{'...' if len(shop_names) > 5 else ''}")
        # Show if any names were split
        split_names = [(orig, tgt) for orig, tgt in zip(shop_names, shop_names_for_targeting) if orig != tgt]
        if split_names:
            print(f"  CL3 targeting (split): {', '.join([f'{tgt} (from {orig})' for orig, tgt in split_names[:3]])}{'...' if len(split_names) > 3 else ''}")

        # Look up deepest_cats for this maincat_id ONCE for the entire group
        deepest_cats = cat_ids_mapping.get(maincat_id_str, [])
        if not deepest_cats:
            print(f"  ⚠️  No deepest_cats found for maincat_id={maincat_id_str}")
            # Mark all rows in this group as failed
            for idx in row_indices:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_ERROR + 1).value = f"No deepest_cats for maincat_id={maincat_id_str}"
                error_count += 1
            continue

        print(f"  Found {len(deepest_cats)} deepest_cat(s)")

        # Track results per shop
        shop_results = {shop: {'success': 0, 'not_found': 0, 'errors': []} for shop in shop_names}
        campaigns_found = 0
        total_exclusions_removed = 0

        # Process each deepest_cat ONCE for all shops in this group
        for deepest_cat in deepest_cats:
            campaign_name = f"PLA/{deepest_cat}_{cl1_str}"
            campaign_data = campaign_cache.get(campaign_name)

            if campaign_data:
                campaigns_found += 1
                ad_groups = campaign_data['ad_groups']

                for ag in ad_groups:
                    ag_id = str(ag['id'])
                    ag_name = ag['name']

                    # Retry logic for connection errors
                    max_retries = 3
                    retry_delay = 2

                    for attempt in range(max_retries):
                        try:
                            # Call batch function with targeting names (split at |)
                            # Use unique targeting names to avoid duplicates
                            unique_targeting_names = list(set(shop_names_for_targeting))
                            result = reverse_exclusion_batch(
                                client=client,
                                customer_id=customer_id,
                                ad_group_id=ag_id,
                                ad_group_name=ag_name,
                                shop_names=unique_targeting_names
                            )

                            # Aggregate results - map targeting names back to original names
                            for targeting_name in result['success']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['success'] += 1
                                        total_exclusions_removed += 1
                            for targeting_name in result['not_found']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['not_found'] += 1
                            for targeting_name, error in result['errors']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['errors'].append(f"{ag_name}: {error}")

                            break  # Success, exit retry loop

                        except Exception as e:
                            error_str = str(e)
                            if "failed to connect" in error_str.lower() or "unavailable" in error_str.lower():
                                if attempt < max_retries - 1:
                                    print(f"    ⚠️  Connection error, retrying in {retry_delay}s...")
                                    time.sleep(retry_delay)
                                    retry_delay *= 2
                                    continue
                            # Non-retryable error or max retries reached
                            error_msg = str(e)[:50]
                            print(f"    ❌ {ag_name}: {error_msg}")
                            for shop in shop_names:
                                shop_results[shop]['errors'].append(f"{ag_name}: {error_msg}")
                            break

                    # Rate limiting delay after each ad group (not each shop!)
                    time.sleep(0.3)

        print(f"  Found {campaigns_found} campaign(s), removed {total_exclusions_removed} exclusion(s) total")

        # =========================================================================
        # STEP 3: Update row statuses based on results
        # =========================================================================
        for idx, shop_name in rows:
            result = shop_results[shop_name]

            # Consider success if: at least one exclusion removed OR not found (wasn't excluded)
            # AND no errors occurred
            has_errors = len(result['errors']) > 0
            has_activity = result['success'] > 0 or result['not_found'] > 0

            if campaigns_found == 0:
                # No campaigns found at all - this is an error
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_ERROR + 1).value = f"No campaigns found for maincat_id={maincat_id_str}"
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ No campaigns")
            elif has_errors:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                error_summary = "; ".join(result['errors'][:3])
                sheet.cell(row=idx, column=COL_ERROR + 1).value = error_summary[:100]
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ {len(result['errors'])} error(s)")
            else:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = True
                sheet.cell(row=idx, column=COL_ERROR + 1).value = ""
                success_count += 1
                print(f"    Row {idx} ({shop_name}): ✅ removed={result['success']}, not_found={result['not_found']}")

        # Save periodically (every N groups)
        if file_path and groups_processed % save_interval == 0:
            print(f"\n💾 Saving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"REVERSE EXCLUSION SHEET SUMMARY (OPTIMIZED)")
    print(f"{'='*70}")
    print(f"Total groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Failed: {error_count + len(rows_with_missing_fields)}")
    print(f"{'='*70}\n")

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """
    Main execution function.
    """
    print(f"\n{'='*70}")
    print("DMA SHOP CAMPAIGNS PROCESSOR")
    print(f"{'='*70}")
    print(f"Operating System: {platform.system()}")
    print(f"Customer ID: {CUSTOMER_ID}")
    print(f"Excel File: {EXCEL_FILE_PATH}")
    print(f"{'='*70}\n")

    # Initialize Google Ads client
    client = initialize_google_ads_client()

    # Create a working copy of the Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    working_copy_path = EXCEL_FILE_PATH.replace(".xlsx", f"_working_copy_{timestamp}.xlsx")

    print(f"\n{'='*70}")
    print(f"CREATING WORKING COPY")
    print(f"{'='*70}")
    print(f"Original file: {EXCEL_FILE_PATH}")
    print(f"Working copy:  {working_copy_path}")

    try:
        shutil.copy2(EXCEL_FILE_PATH, working_copy_path)
        print(f"✅ Working copy created successfully")
    except Exception as e:
        print(f"❌ Error creating working copy: {e}")
        sys.exit(1)

    # Load Excel workbook from working copy
    print(f"\n{'='*70}")
    print(f"LOADING WORKING COPY")
    print(f"{'='*70}")
    print(f"Loading: {working_copy_path}")
    try:
        workbook = load_workbook(working_copy_path)
        print(f"✅ Excel file loaded successfully")
        print(f"   Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        sys.exit(1)
    ''' 
    # Process exclusion sheet (V2 - with cat_ids mapping)
    try:
        process_exclusion_sheet_v2(client, workbook, CUSTOMER_ID, working_copy_path)
    except Exception as e:
        print(f"❌ Error processing exclusion sheet: {e}")

    # Load reverse exclusion file separately
    print(f"\n{'='*70}")
    print(f"LOADING REVERSE EXCLUSION FILE")
    print(f"{'='*70}")
    print(f"File: {REVERSE_EXCLUSION_FILE_PATH}")
    '''

    reverse_working_copy_path = REVERSE_EXCLUSION_FILE_PATH.replace(".xlsx", f"_working_copy_{timestamp}.xlsx")

    try:
        shutil.copy2(REVERSE_EXCLUSION_FILE_PATH, reverse_working_copy_path)
        print(f"✅ Reverse exclusion working copy created: {reverse_working_copy_path}")
        reverse_workbook = load_workbook(reverse_working_copy_path)
        print(f"✅ Reverse exclusion file loaded successfully")
        print(f"   Available sheets: {reverse_workbook.sheetnames}")

        #process_reverse_exclusion_sheet(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path, 10, "verwijderen")
        #process_reverse_inclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        #process_enable_inclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path, "toevoegen")
        process_exclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        #process_inclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)

        # Save reverse exclusion workbook
        reverse_workbook.save(reverse_working_copy_path)
        print(f"✅ Reverse exclusion results saved to: {reverse_working_copy_path}")
    except FileNotFoundError:
        print(f"⚠️  Reverse exclusion file not found: {REVERSE_EXCLUSION_FILE_PATH}")
        print(f"   Skipping reverse exclusion processing")
    except Exception as e:
        print(f"❌ Error processing reverse exclusion file: {e}")

    '''
    # Validate cl1 targeting (Dry run)
    validate_cl1_targeting_for_campaigns(client, CUSTOMER_ID, "% store_%", False)    
    
    # Process inclusion sheet (V2 - new structure)
    try:
        process_inclusion_sheet_legacy(client, workbook, CUSTOMER_ID, working_copy_path)
    except Exception as e:
        print(f"❌ Error processing inclusion sheet: {e}")
        
    # Process uitbreiding sheet
    try:
       process_uitbreiding_sheet(client, workbook, CUSTOMER_ID, working_copy_path)
    except Exception as e:
       print(f"❌ Error processing uitbreiding sheet: {e}")
    '''

    # Final save to working copy
    print(f"\n{'='*70}")
    print("SAVING FINAL RESULTS")
    print(f"{'='*70}")
    print(f"All results saved to working copy: {working_copy_path}")
    print(f"Original file remains unchanged: {EXCEL_FILE_PATH}")
    print(f"\nTo use the results, rename or copy the working copy to:")
    print(f"  {EXCEL_FILE_PATH}")
    print(f"{'='*70}")

    print(f"\n{'='*70}")
    print("PROCESSING COMPLETE")
    print(f"{'='*70}\n")

if __name__ == "__main__":
    main()
