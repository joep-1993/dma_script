"""
Test script to verify improved migration with:
1. Incremental saving
2. Rate limiting
3. Proper error handling (no false SUCCESS on CONCURRENT_MODIFICATION)

This processes a small batch of 10 campaigns as a test.
"""

import sys
from openpyxl import load_workbook
from campaign_processor import (
    initialize_google_ads_client,
    process_exclusion_sheet,
    EXCEL_FILE_PATH,
    SHEET_EXCLUSION,
    COL_EX_STATUS
)


def test_improved_migration(customer_id: str, start_row: int = 26, end_row: int = 35):
    """
    Test the improved migration on a small batch.

    Args:
        customer_id: Google Ads customer ID
        start_row: First row to process (inclusive)
        end_row: Last row to process (inclusive)
    """
    print("="*70)
    print("TESTING IMPROVED MIGRATION")
    print("="*70)
    print(f"Customer ID: {customer_id}")
    print(f"Test range: Rows {start_row} to {end_row} ({end_row - start_row + 1} campaigns)")
    print(f"Excel file: {EXCEL_FILE_PATH}")
    print("="*70)

    # Load workbook
    print("\nLoading Excel file...")
    workbook = load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook[SHEET_EXCLUSION]

    # Clear status for test rows
    print(f"\nClearing status for rows {start_row} to {end_row}...")
    cleared_count = 0
    for row_num in range(start_row, end_row + 1):
        row = sheet[row_num]
        if row[COL_EX_STATUS].value is not None:
            row[COL_EX_STATUS].value = None
            cleared_count += 1

    # Save cleared status
    workbook.save(EXCEL_FILE_PATH)
    print(f"✅ Cleared {cleared_count} row(s)")

    # Initialize Google Ads client
    print("\nInitializing Google Ads client...")
    client = initialize_google_ads_client()

    # Reload workbook for processing
    workbook = load_workbook(filename=EXCEL_FILE_PATH)

    # Process exclusion sheet with improved parameters
    print("\nProcessing exclusion sheet with improvements:")
    print("  - Save interval: Every 5 campaigns (small batch)")
    print("  - Rate limit: 0.5s delay between campaigns")
    print()

    process_exclusion_sheet(
        client=client,
        workbook=workbook,
        customer_id=customer_id,
        save_interval=5,  # Save every 5 campaigns for testing
        rate_limit_seconds=0.5  # 500ms delay between campaigns
    )

    # Close workbook
    workbook.close()

    # Check results
    print("\n" + "="*70)
    print("CHECKING RESULTS")
    print("="*70)

    workbook = load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook[SHEET_EXCLUSION]

    success_count = 0
    failure_count = 0

    print(f"\nResults for rows {start_row} to {end_row}:")
    for row_num in range(start_row, end_row + 1):
        row = sheet[row_num]
        shop_name = row[0].value  # COL_EX_SHOP_NAME (Column A)
        category = row[2].value   # COL_EX_CAT_UITSLUITEN (Column C)
        cl1 = row[4].value        # COL_EX_CUSTOM_LABEL_1 (Column E)
        status = row[COL_EX_STATUS].value

        if status == True:
            success_count += 1
            status_str = "✅ SUCCESS"
        elif status == False:
            failure_count += 1
            status_str = "❌ FAILED"
        else:
            status_str = "⚠️  UNKNOWN"

        print(f"  Row {row_num}: {status_str} - {shop_name} | {category}_{cl1}")

    workbook.close()

    print(f"\n" + "="*70)
    print(f"TEST SUMMARY")
    print(f"="*70)
    print(f"Total processed: {success_count + failure_count}")
    print(f"Successful: {success_count}")
    print(f"Failed: {failure_count}")
    print(f"="*70)


def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python3 test_improved_migration.py <customer_id> [start_row] [end_row]")
        print("\nExample:")
        print("  python3 test_improved_migration.py 3800751597")
        print("  python3 test_improved_migration.py 3800751597 26 35")
        print("\nDefault: Rows 26-35 (10 campaigns)")
        sys.exit(1)

    customer_id = sys.argv[1]
    start_row = int(sys.argv[2]) if len(sys.argv) >= 3 else 26
    end_row = int(sys.argv[3]) if len(sys.argv) >= 4 else 35

    test_improved_migration(customer_id, start_row, end_row)


if __name__ == "__main__":
    main()
